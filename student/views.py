# Create your views here.
import os

import openpyxl
from django.db.models import Q, Sum
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.forms.models import model_to_dict
from django.db import IntegrityError, transaction
from django.conf import settings
from fpdf import FPDF
from fpdf.enums import VAlign, TableCellFillMode, TableHeadingsDisplay
from fpdf.table import Table
from openpyxl.utils.datetime import to_excel
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from authentification.models import SchoolYear, User
from note.models import Note, Enseignements
from classroom.models import ClassRoom
from note.views import ReportCard
from .forms import StudentForm, ParentForm, DForm, BulkDecisionForm
from osm.forms import SearchForm
from osm.forms import SearchForm
from note.forms import CheckForm, MarksForm, SelectForm
from .models import Parent, Student, StudentDiscipline, EnrollmentStatus, StudentEnrollment
from osm.utils import formated_float, message, logged_admin_view, LoggedAdminView, ListView, DeleteView, ADetailView, \
    with_users_school_schema, school_year, pdf_response, resize_image, LoggedAdminOrTitulaireView, zip_pdfs_response, \
    check_notes, stamp_bytes, paste_stamp, base_header, safe_redirect_back, filigrane, add_fonts, format_date
from pandas import DataFrame, read_excel, ExcelWriter, isnull, Timestamp, to_datetime
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.styles import Alignment, Font
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime, date
from os import path


"""
=============================================================================
 VIEW — Téléchargement du MODÈLE d'import avec liste déroulante des classes
=============================================================================
Principe :
  - on lit le modèle statique (static/document/Modèle_Liste_des_Élèves.xlsx) ;
  - on injecte une VALIDATION DE DONNÉES (liste déroulante Excel) sur la
    colonne Classe (G), alimentée par les classes RÉELLES de l'établissement
    courant + une option "(Aucune)" ;
  - on renvoie le fichier en téléchargement (BytesIO -> HttpResponse).

Structure du modèle (vérifiée) :
  - feuille "Élèves"
  - en-têtes en ligne 4 ; données à partir de la ligne 5
  - colonne Classe = G  (Statut = H)
"""

# Emplacement du modèle dans les statiques.
TEMPLATE_FILENAME = "document/Modèle Liste des Élèves.xlsx"

HEADER_ROW = 4          # ligne des en-têtes
FIRST_DATA_ROW = 5      # première ligne de données
DATE_COL = "D"          # colonne Date de naissance
CLASS_COL = "G"         # colonne Classe
STATUT_COL = "H"        # colonne Statut
LAST_DATA_ROW = 1000    # jusqu'où appliquer les listes déroulantes

# Largeurs "optimales" par colonne (optionnel : appliquées seulement si
# APPLY_WIDTHS = True).
APPLY_WIDTHS = False
COLUMN_WIDTHS = {
    "A": 13,   # Matricule
    "B": 28,   # Noms
    "C": 19,   # Prénoms
    "D": 17,   # Date de naissance
    "E": 20,   # Lieu de naissance
    "F": 13,   # Sexe
    "G": 16,   # Classe
    "H": 16,   # Statut
}

# Insérer le nom de l'établissement dans le titre (optionnel).
SHOW_SCHOOL_NAME = True


def _find_template_path():
    """
    Localise le modèle dans les répertoires statiques. On essaie d'abord
    finders (collectstatic non requis en dev), puis STATIC_ROOT.
    """
    # 1) via les finders
    try:
        from django.contrib.staticfiles import finders
        p = finders.find(TEMPLATE_FILENAME)
        if p:
            return p
    except Exception:
        pass
    # 2) repli : STATIC_ROOT
    if getattr(settings, "STATIC_ROOT", None):
        p = os.path.join(settings.STATIC_ROOT, TEMPLATE_FILENAME)
        if os.path.exists(p):
            return p
    return None


def _make_list_validation(wb, ws, options, sheet_name, anchor):
    """
    Construit une DataValidation de type liste pour 'options'. Si la liste est
    courte (et sans virgule), on l'inline ; sinon on l'écrit dans une feuille
    cachée dédiée (sheet_name) et on pointe dessus. Renvoie la DataValidation.
    """
    joined = ",".join(options)
    if len(joined) <= 250 and not any("," in (o or "") for o in options):
        dv = DataValidation(type="list", formula1=f'"{joined}"',
                            allow_blank=True, showDropDown=False)
    else:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        src = wb.create_sheet(sheet_name)
        for i, val in enumerate(options, start=1):
            src.cell(row=i, column=1, value=val)
        src.sheet_state = "hidden"
        ref = f"{quote_sheetname(sheet_name)}!$A$1:$A${len(options)}"
        dv = DataValidation(type="list", formula1=ref,
                            allow_blank=True, showDropDown=False)
    return dv


class ImportTemplateDownload(LoggedAdminView):
    """
    Génère et renvoie le modèle d'import avec :
      - colonne Classe (G) en liste déroulante (classes de l'établissement + (Aucune)) ;
      - colonne Statut (H) en liste déroulante (Nouveau / Redoublant) ;
      - (option) largeurs de colonnes ;
      - (option) nom de l'établissement dans le titre.
    """

    def get(self, *args, **kwargs):
        path = _find_template_path()
        if not path:
            return HttpResponse("Modèle introuvable sur le serveur.", status=500)

        wb = openpyxl.load_workbook(path)
        ws = wb["Élèves"] if "Élèves" in wb.sheetnames else wb.active

        # --- 1) Liste déroulante CLASSE (G) ---
        codes = list(
            ClassRoom.objects.order_by_niveau().values_list("code", flat=True)
        )
        class_options = ["(Aucune)"] + [c for c in codes if c]
        dv_class = _make_list_validation(wb, ws, class_options, "_classes", CLASS_COL)
        dv_class.error = "Choisissez une classe dans la liste (ou (Aucune))."
        dv_class.errorTitle = "Classe invalide"
        dv_class.prompt = "Sélectionnez la classe de l'élève."
        dv_class.promptTitle = "Classe"
        ws.add_data_validation(dv_class)
        dv_class.add(f"{CLASS_COL}{FIRST_DATA_ROW}:{CLASS_COL}{LAST_DATA_ROW}")

        # --- 2) Liste déroulante STATUT (H) ---
        statut_options = ["Nouveau", "Redoublant"]
        dv_statut = _make_list_validation(wb, ws, statut_options, "_statuts", STATUT_COL)
        dv_statut.error = "Choisissez « Nouveau » ou « Redoublant » (ou laissez vide)."
        dv_statut.errorTitle = "Statut invalide"
        dv_statut.prompt = "Nouveau ou Redoublant (laisser vide = Nouveau par défaut)."
        dv_statut.promptTitle = "Statut"
        ws.add_data_validation(dv_statut)
        dv_statut.add(f"{STATUT_COL}{FIRST_DATA_ROW}:{STATUT_COL}{LAST_DATA_ROW}")

        # --- 2bis) Colonne DATE DE NAISSANCE (D) : format forcé jj/mm/aaaa ---
        # Double protection contre l'inversion jour/mois :
        #  (a) on impose le FORMAT D'AFFICHAGE jj/mm/aaaa sur les cellules ->
        #      Excel range une vraie date non ambiguë ;
        #  (b) on ajoute une validation "date" qui borne les années plausibles,
        #      ce qui force aussi Excel à traiter la saisie comme une date.
        from openpyxl.styles import Alignment as _Align
        from datetime import date as _date
        for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            cell = ws[f"{DATE_COL}{r}"]
            cell.number_format = "DD/MM/YYYY"
            cell.alignment = _Align(horizontal="center")

        now_year = __import__("datetime").datetime.now().year
        min_year, max_year = now_year - 30, now_year - 8
        dv_date = DataValidation(
            type="date", operator="between",
            formula1=int(to_excel(_date(min_year, 1, 1))),
            formula2=int(to_excel(_date(max_year, 12, 31))),
            allow_blank=True, showErrorMessage=True,
        )
        dv_date.error = f"Entrez une date de naissance valide au format jj/mm/aaaa (année comprise entre {min_year} et {max_year} (inclus)"
        dv_date.errorTitle = "Date invalide"
        dv_date.prompt = f"Format : jj/mm/aaaa (ex. 15/03/{now_year-12})."
        dv_date.promptTitle = "Date de naissance"
        ws.add_data_validation(dv_date)
        dv_date.add(f"{DATE_COL}{FIRST_DATA_ROW}:{DATE_COL}{LAST_DATA_ROW}")

        # --- 3) (Option) Largeurs de colonnes ---
        if APPLY_WIDTHS:
            for col, width in COLUMN_WIDTHS.items():
                ws.column_dimensions[col].width = width

        # --- 4) (Option) Nom de l'établissement dans le titre fusionné (A2) ---
        if SHOW_SCHOOL_NAME:
            try:
                school = getattr(self.request.user, "school", None)
                nom_etab = getattr(school, "nom", None) if school else None
                if nom_etab:
                    # A2 est la cellule "ancre" de la fusion A2:H3 -> on y écrit.
                    ws["A2"] = f"Liste des Élèves — {nom_etab}"
            except Exception:
                pass   # purement cosmétique : on n'échoue jamais là-dessus

        # --- Sérialisation + téléchargement ---
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        resp = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="Modele Liste des Eleves.xlsx"'
        return resp


"""
=============================================================================
 VIEW — Attribution de classe en masse (aux élèves sans classe)
=============================================================================
Mécanisme:
  - un SELECT de classe par ligne (cas isolés / hétérogènes) ;
  - une barre "sélection + classe groupée" qui PRÉ-REMPLIT (côté JS) le select
    des lignes cochées ; l'utilisateur peut ensuite ajuster chaque ligne ;
  - à l'enregistrement, le serveur ne lit QUE le select de chaque ligne
    (classe_<student_id>). Une seule logique, aucune ambiguïté.

On fait des save() individuels (et pas bulk_update) :
  affecter une classe déclenche le signal post_save -> création automatique du
  StudentEnrollment de l'année courante. On VEUT ce signal ici, on fait save() par
  élève, dans une transaction. On prévoit que les volumes (sans-classe) sont faibles.
"""


class AssignClassToWithoutClass(LoggedAdminView):
    template_name = "class_bulk_assignment.html"

    def get_students(self):
        return Student.objects.filter(classe__isnull=True).order_by_classroom_level()

    def get(self, *args, **kwargs):
        students = self.get_students()
        context = {
            "title": "Attribuer une classe (aux élèves sans classe)",
            "students": students,
            "classrooms": ClassRoom.objects.select_related("classe").order_by_niveau(),
        }
        return render(self.request, self.template_name, context)

    def post(self, *args, **kwargs):
        students = list(Student.objects.filter(classe__isnull=True))

        # Ids de classe valides (sécurité).
        valid_class_ids = set(ClassRoom.objects.values_list("id", flat=True))

        assigned = 0
        with transaction.atomic():
            for student in students:
                raw = self.request.POST.get(f"classe_{student.id}") or None
                if not raw:
                    continue                       # ligne laissée vide -> on ignore
                try:
                    cid = int(raw)
                except (TypeError, ValueError):
                    continue
                if cid not in valid_class_ids:
                    continue

                student.classe_id = cid
                # save() individuel -> déclenche le signal post_save qui crée le
                # StudentEnrollment de l'année courante s'il n'existe pas.
                student.save(update_fields=["classe"])
                assigned += 1

        if assigned:
            message(self.request, f"{assigned} élève(s) affecté(s) à une classe.")
        else:
            message(self.request, "Aucune affectation effectuée.", msg_type="warning")
        return redirect("students_without_class")


"""
=============================================================================
 VIEWS — Corbeille (désactivés) & À régulariser (actifs sans classe)
=============================================================================
"""
# -----------------------------------------------------------------------------
# Base commune aux deux écrans secondaires : on réutilise students_list.html.
# -----------------------------------------------------------------------------
class _StudentSubListView(LoggedAdminView):
    """
    Classe de base : rend students_list.html avec un queryset et un contexte
    spécifiques. Les sous-classes définissent get_queryset() + les libellés.
    """
    template_name = "students_list.html"
    title = ""
    info = ""
    review_mode = None          # "trash" | "without_class" -> pilote le bouton groupé

    def get_queryset(self):
        raise NotImplementedError

    def get(self, *args, **kwargs):
        datas = self.get_queryset()
        context = {
            "title": self.title,
            "info": self.info,
            "datas": datas,
            "review_mode": self.review_mode,   # le template affiche le bon bouton
            "pk": None, # pas de pk -> students_list.html sait qu'on n'est pas dans une classe
            # on masque la barre de recherche/ajout sur ces écrans secondaires
            "secondary": True,
        }
        return render(self.request, self.template_name, context)


# -----------------------------------------------------------------------------
# Corbeille : les élèves désactivés
# -----------------------------------------------------------------------------
class StudentsTrash(_StudentSubListView):
    title = "Corbeille — élèves désactivés"
    info = "Élèves désactivés. Vous pouvez les réactiver, ou les supprimer définitivement."
    review_mode = "trash"

    def get_queryset(self):
        return Student.objects_all.filter(is_active=False).order_by_classroom_level()


# -----------------------------------------------------------------------------
# À régulariser : les actifs sans classe
# -----------------------------------------------------------------------------
class StudentsWithoutClass(_StudentSubListView):
    title = "À régulariser — élèves sans classe"
    info = "Élèves actifs non affectés à une classe. Attribuez-leur une classe, ou supprimez-les."
    review_mode = "without_class"

    def get_queryset(self):
        # Actifs (objects) + sans classe.
        return Student.objects.filter(classe__isnull=True).order_by_classroom_level()


# -----------------------------------------------------------------------------
# Toggle activer/désactiver (unitaire)
# -----------------------------------------------------------------------------
class StudentToggleActive(LoggedAdminView):
    def post(self, *args, **kwargs):
        student = get_object_or_404(Student.objects_all, pk=self.kwargs["id"])
        if student.is_active:
            student.deactivate()
            message(self.request, f"{student} a été désactivé(e) et n'apparaîtra plus dans les listes, bulletins et autres.")
        else:
            student.activate()
            message(self.request, f"{student} a été réactivé(e).")
        nxt = self.request.POST.get("next")
        return redirect(nxt) if nxt else redirect("students")


# -----------------------------------------------------------------------------
# Suppression groupée — désactivés (depuis la corbeille)
# -----------------------------------------------------------------------------
class DeleteDeactivatedStudents(LoggedAdminView):
    template_name = "students_bulk_delete.html"

    def get_targets(self):
        return Student.objects_all.filter(is_active=False)

    def get(self, *args, **kwargs):
        targets = self.get_targets()
        return render(self.request, self.template_name, {
            "title": "Vider la corbeille",
            "count": targets.count(),
            "alerte": ("Cette action est IRRÉVERSIBLE : tous les élèves désactivés "
                       "seront définitivement supprimés (notes, parcours, discipline, photos)."),
            "back": "students_trash",
        })

    def post(self, *args, **kwargs):
        targets = self.get_targets()
        n = targets.count()
        if n:
            with transaction.atomic():
                for student in targets:
                    student.delete()
            message(self.request, f"{n} élève(s) désactivé(s) supprimé(s) définitivement.")
        else:
            message(self.request, "La corbeille est déjà vide.", msg_type="warning")
        return redirect("students_trash")


# -----------------------------------------------------------------------------
# Suppression groupée — actifs sans classe (depuis l'écran à régulariser)
# -----------------------------------------------------------------------------
class DeleteStudentsWithoutClass(LoggedAdminView):
    template_name = "students_bulk_delete.html"

    def get_targets(self):
        return Student.objects.filter(classe__isnull=True)   # actifs sans classe

    def get(self, *args, **kwargs):
        targets = self.get_targets()
        return render(self.request, self.template_name, {
            "title": "Supprimer les élèves sans classe",
            "count": targets.count(),
            "alerte": ("Cette action est IRRÉVERSIBLE : notes, parcours, discipline "
                       "et photos de ces élèves seront définitivement perdus. "
                       "Pour seulement les retirer des listes, désactivez-les plutôt."),
            "back": "students_without_class",
        })

    def post(self, *args, **kwargs):
        targets = self.get_targets()
        n = targets.count()
        if n:
            with transaction.atomic():
                for student in targets:
                    student.delete()
            message(self.request, f"{n} élève(s) sans classe supprimé(s) définitivement.")
        else:
            message(self.request, "Aucun élève sans classe à supprimer.", msg_type="warning")
        return redirect("students_without_class")


"""
=============================================================================
 VIEW — Parcours scolaire d'un élève au sein de l'établissement
=============================================================================
"""


class StudentJourney(LoggedAdminView):
    template_name = "student_journey.html"

    def get(self, *args, **kwargs):
        # Élève + relations affichées dans l'en-tête (classe actuelle, parents).
        student = get_object_or_404(
            Student.objects_all.select_related("classe", "pere", "mere"),
            pk=self.kwargs["id"]
        )

        # Historique : un enrollment par année, le plus récent en premier.
        # select_related évite tout N+1 dans le tableau (année, classe, classe
        # de destination sont lus pour chaque ligne).
        enrollments = (
            student.enrollments
            .select_related("school_year", "classroom", "next_classroom")
            .order_by("-school_year__annee_debut")
        )

        context = {
            "title": f"Parcours — {student}",
            "student": student,
            "enrollments": enrollments,
            "nb_annees": enrollments.count(),
        }
        return render(self.request, self.template_name, context)


# -----------------------------------------------------------------------------
# Tableau formulaire d'attribution des décisions de fin d'année
# -----------------------------------------------------------------------------
class EndYearAssignmentForm(LoggedAdminOrTitulaireView):
    template_name = "end_year_assignment.html"

    # --- Helpers ---------------------------------------------------------------
    def get_classroom(self, method="GET"):
        """Charge la classe ciblée (+ relations utiles)."""
        return get_object_or_404(
            ClassRoom.objects.select_related("classe", "titulaire"),
            pk=int(self.request.GET['classroom'] if method == "GET" else self.request.POST['classroom'])
        )

    def build_bulk_form(self, classroom):
        """
        Instancie le BulkDecisionForm UNE fois. Il porte promote_qs / repeat_qs
        (les deux listes de classes générées une seule fois côté form).
        """
        return BulkDecisionForm(context={
            "request": self.request,
            "classroom": classroom,
        })

    def get_context(self, classroom, bulk_form):
        current_year = SchoolYear.current()

        # Élèves de la classe, ordonnés.
        students = classroom.students.all().order_by("nom", "prenom")

        # Enrollments de l'année courante chargés en UNE requête, indexés par
        # student_id (évite tout N+1 dans la boucle ci-dessous).
        enrollments = {}
        if current_year:
            qs = StudentEnrollment.objects.filter(
                school_year=current_year, student__in=students
            ).select_related("next_classroom")
            enrollments = {e.student_id: e for e in qs}

        # Lignes du tableau : valeurs déjà saisies pré-sélectionnées.
        rows = []
        for student in students:
            enr = enrollments.get(student.id)
            rows.append({
                "student": student,
                "enrollment": enr,
                "decision": enr.decision if enr else EnrollmentStatus.EN_COURS,
                "next_classroom_id": enr.next_classroom_id if enr else None,
            })

        # Décisions proposées (on retire "En cours").
        decision_choices = [
            (value, label) for value, label in EnrollmentStatus.choices
            if value not in [EnrollmentStatus.EN_COURS, EnrollmentStatus.NON_STATUE]
        ]

        return {
            "title": f"Décisions de fin d'année — {classroom.code}",
            "classroom": classroom,
            "rows": rows,
            "decision_choices": decision_choices,
            "current_year": current_year,
            "bulk_form": bulk_form,                  # porte promote_qs / repeat_qs
            # Valeurs injectées au JS pour choisir la bonne liste de classes :
            "promote_value": EnrollmentStatus.PROMU,
            "repeat_value": EnrollmentStatus.REDOUBLE,
        }

    def get(self, *args, **kwargs):
        classroom = self.get_classroom()
        bulk_form = self.build_bulk_form(classroom)
        context = self.get_context(classroom, bulk_form)
        return render(self.request, self.template_name, context)

    def post(self, *args, **kwargs):
        classroom = self.get_classroom(method="POST")
        current_year = SchoolYear.current()

        if current_year is None:
            message(self.request,
                    "Aucune année scolaire courante définie. Impossible d'enregistrer.",
                    msg_type="error")
            return redirect("classrooms")

        students = list(classroom.students.all())

        # Enrollments existants pour cette classe/année, indexés par student_id.
        existing = {
            e.student_id: e
            for e in StudentEnrollment.objects.filter(
                school_year=current_year, student__in=students
            )
        }

        # Sécurité serveur : n'accepter une classe cible que si elle
        # appartient à la bonne liste selon la décision. On réutilise le MÊME
        # bulk_form (donc les listes déjà générées, pas de requête en plus).
        bulk_form = self.build_bulk_form(classroom)
        promote_ids = {c.id for c in bulk_form.promote_qs}
        repeat_ids = {c.id for c in bulk_form.repeat_qs}

        decided_by = self.request.user.staff_member.short_name or self.request.user.username
        now = timezone.now()

        to_update, to_create, nb = [], [], 0

        for student in students:
            decision = self.request.POST.get(f"decision_{student.id}")
            next_id = self.request.POST.get(f"next_classroom_{student.id}") or None

            # Ligne non renseignée -> ignorée.
            if not decision:
                continue

            # Décisions sans classe cible -> on neutralise next_id.
            if decision in (EnrollmentStatus.TRANSFERE, EnrollmentStatus.SORTI, EnrollmentStatus.EXCLU):
                next_id = None
            # Décisions avec classe cible -> on vérifie la cohérence de la liste.
            elif decision == EnrollmentStatus.PROMU:
                if next_id and int(next_id) not in promote_ids:
                    next_id = None
            elif decision == EnrollmentStatus.REDOUBLE:
                if next_id and int(next_id) not in repeat_ids:
                    next_id = None

            next_pk = int(next_id) if next_id else None

            enr = existing.get(student.id)
            if enr:
                changed = (
                    enr.decision != decision
                    or (enr.next_classroom_id or None) != next_pk
                )
                if changed:
                    enr.decision = decision
                    enr.next_classroom_id = next_pk
                    enr.decided_by = decided_by
                    enr.decided_at = now
                    to_update.append(enr)
                    nb += 1
            else:
                # Cas rare : élève sans enrollment courant.
                to_create.append(StudentEnrollment(
                    student=student,
                    school_year=current_year,
                    classroom=classroom,
                    decision=decision,
                    next_classroom_id=next_pk,
                    decided_by=decided_by,
                    decided_at=now,
                ))
                nb += 1

        if to_update:
            StudentEnrollment.objects.bulk_update(
                to_update,
                ["decision", "next_classroom_id", "decided_by", "decided_at"]
            )
        if to_create:
            StudentEnrollment.objects.bulk_create(to_create)

        if nb:
            message(self.request,
                    f"Décisions de fin d'année enregistrées pour {nb} élève(s) en {classroom.code}.")
        else:
            message(self.request, "Aucune modification effectuée.", msg_type="warning")

        response = render(
            self.request,
            self.template_name,
            self.get_context(classroom,
                             self.build_bulk_form(classroom))
        )
        response['HX-Trigger'] = 'AJAXMessages'
        return response


class EndYearAssignment(LoggedAdminOrTitulaireView):
    template_name = "edit_marks.html"
    title = "Attribution des classe pour l'année prochaine"

    def get(self, *args, **kwargs):
        select_form = SelectForm(context={"request": self.request, 'trim': False, 'marks_sheet': True,
                                          'enseignements': None, 'end_year_assignment': True})
        context = {'end_year_assignment': True, 'title': self.title, 'select_form': select_form}
        return render(self.request, self.template_name, context)

    def post(self, *args, **kwargs):
        select_form = SelectForm(context={"request": self.request, 'trim': False, 'marks_sheet': True,
                                          'enseignements': None, 'end_year_assignment': True})
        context = {'end_year_assignment': True, 'title': self.title, 'select_form': select_form}
        return render(self.request, self.template_name, context)


class StudentsIdCards(LoggedAdminView):
    template_name = "students_id_cards.html"
    title = "Cartes d'Identité Scolaire"

    def get(self, *args, **kwargs):
        select_form = SelectForm(context={
            "request": self.request, 'trim': False, 'marks_sheet': True, 'enseignements': None})
        context = {'marks_sheet': True, 'csi': True, 'title': self.title, 'select_form': select_form}
        return render(self.request, self.template_name, context)

    @staticmethod
    def build_pdf_or_reason(classroom, data):
        reason = check_notes(classroom, None, marks_sheet=True)
        if reason is not None:
            return reason  # -> sautée (ZIP) ou message d'erreur (une classe)
        data['students'] = list(classroom.students.order_by('nom', 'prenom'))
        return StudentsIdentityCards(data=data) if data['layout'] == "fold" else StudentsIdentityCardsCNI(data=data)

    def post(self, *args, **kwargs):
        empty_csi = True if 'csi_checkbox' in self.request.POST.keys() else False
        annee = school_year()
        school = User.objects.select_related('school').get(id=self.request.user.id).school
        data = {'annee': annee, 'school_data': school.school_to_dict()}
        selected = self.request.POST.get("classroom")
        filename = self.title
        if not empty_csi:
            data['layout'] = self.request.POST.get('layout', 'fold')
            # -------- Cas "Toutes les classes" -> ZIP --------
            if selected == "__all__":
                classrooms = (
                    ClassRoom.objects.prefetch_related('students__pere', 'students__mere')
                )

                def build(clsrm):
                    return self.build_pdf_or_reason(clsrm, data)

                def namer(clsrm):
                    return f"{clsrm.code} {filename}.pdf"

                return zip_pdfs_response(
                    build_pdf_for_classroom=build,
                    classrooms=classrooms,
                    zip_filename=f"Cartes dIdentité Scolaire - Toutes les classes.zip",
                    per_file_namer=namer,
                )

            # -------- Cas "une seule classe" --------
            classroom = (
                ClassRoom.objects.prefetch_related('students__pere', 'students__mere').
                get(pk=int(selected))
            )
            result = self.build_pdf_or_reason(classroom, data)
            if isinstance(result, str):
                return JsonResponse({'success': False, 'message': result})
            return pdf_response(result, f"{filename} {classroom.code}.pdf")
        return pdf_response(StudentsIdentityCards(data=data), f"{filename}.pdf")


# Exportation de la liste des élèves dans un fichier Excel
@logged_admin_view
def students_export(request, cls_id):
    queryset = Student.objects.select_related('classe').order_by_classroom_level()
    classroom = None
    columns = ['Matricule', 'Noms', 'Prénoms', 'Date de naissance', 'Lieu de Naissance', 'Sexe', 'Statut']
    if cls_id:
        classroom = ClassRoom.objects.get(pk=cls_id)
        students = (
            queryset.filter(classe_id=classroom.pk)
            .values('unique_id', 'nom', "prenom", 'date_naissance', 'lieu_naissance', 'sexe', 'statut')
        )
        details = f"Effecif : {classroom.effectif}, {classroom.kind_numbers}"
    else:
        students = (
            queryset.values('unique_id', 'nom', "prenom", 'date_naissance', 'lieu_naissance', 'sexe', 'classe__code',
                            'statut')
        )
        columns.insert(-1, 'Classe')
        details = f"Effectif : {request.user.school.effectif}, {request.user.school.kind_numbers}"
    data_frame = DataFrame(students)
    data_frame.columns = columns
    buffer  = BytesIO()
    with ExcelWriter(buffer, engine='openpyxl') as writer:
        data_frame.to_excel(writer, index=False, startrow=3, sheet_name="Élèves")

    buffer.seek(0)
    workbook = load_workbook(buffer)
    worksheet = workbook.active
    title = "Liste des élèves"
    center = Alignment(horizontal='center', vertical='center')
    if cls_id:
        title += f" de {classroom.code}"
        worksheet.merge_cells('A1:G1')
        worksheet.merge_cells('A2:G2')
        worksheet.merge_cells('A3:G3')
    else:
        worksheet.merge_cells('A1:H1')
        worksheet.merge_cells('A2:H2')
        worksheet.merge_cells('A3:H3')
        worksheet.column_dimensions['H'].width = 10
    worksheet['A1'] = f"{request.user.school.nom} / {request.user.school.name}"
    worksheet['A1'].font = Font(size=15, bold=True)
    worksheet['A1'].alignment = center
    worksheet.row_dimensions[1].height = 30
    worksheet['A2'] = title
    worksheet['A2'].font = Font(size=13, italic=True)
    worksheet['A2'].alignment = center
    worksheet.row_dimensions[2].height = 20
    worksheet['A3'] = details
    worksheet['A3'].font = Font(size=12, italic=True)
    worksheet['A3'].alignment = center

    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 8
    worksheet.column_dimensions['G'].width = 12

    for col_mumber in range(1, len(data_frame.columns)+1):
        col_letter = get_column_letter(col_mumber)
        cell = worksheet[f"{col_letter}4"]
        cell.font = Font(bold=True)
        cell.alignment = center

    final_buffer = BytesIO()
    workbook.save(final_buffer)
    final_buffer.seek(0)

    response = HttpResponse(final_buffer.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={title}.xlsx'
    return response


# Importation d'une liste d'élèves à partir d'un fichier excel
class StudentsImport(LoggedAdminView):
    title = "Importation d'une liste d'Élèves"
    template_name = "students_import.html"
    required_fiels = ['matricule', 'noms', 'prénoms', 'date de naissance', 'lieu de naissance', 'sexe']

    def get(self, *args, **kwargs):
        return render(self.request, self.template_name, context={'title': self.title})

    def post(self, *args, **kwargs):
        rapport = list
        file = self.request.FILES.get('file')
        if file:
            rapport = self.import_students(file)
        return render(self.request, self.template_name, context={'title': self.title, 'rapport': rapport})

    def parse_birthdate(self, value):
        from datetime import date, datetime as _dt
        """
        Renvoie un datetime.date, ou None si invalide.
        - Si Excel a déjà fourni une date/datetime (ou un Timestamp pandas), on la
          prend telle quelle (NON ambiguë).
        - Si c'est une chaîne, on force le format français jj/mm/aaaa (dayfirst),
          en acceptant les séparateurs / - . et l'année sur 2 ou 4 chiffres.
        """

        if value is None or isnull(value):
            return None

        # 1) Déjà une date/datetime/Timestamp -> Excel l'a stockée comme vraie date.
        if isinstance(value, (Timestamp, _dt)):
            return value.date()
        if isinstance(value, date):
            return value

        # 2) Chaîne -> on force le jour en premier (jj/mm/aaaa), formats explicites.
        s = str(value).strip()
        if not s:
            return None

        # Formats explicites prioritaires (sans ambiguïté).
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%y", "%d-%m-%y"):
            try:
                return _dt.strptime(s, fmt).date()
            except ValueError:
                continue

        # Repli : pandas avec dayfirst=True (jour d'abord), jamais le défaut US.
        dt = to_datetime(s, errors="coerce", dayfirst=True)
        if dt is None or (hasattr(dt, "isnull") and dt.isnull()):
            return None
        try:
            return dt.date()
        except Exception:
            return None

    # Détecter la ligne des en-têtes dans la 10 premières lignes maximum
    def detect_header_row(self, df: DataFrame):
        for i in range(min(10, len(df))):
            row = df.iloc[i].fillna('').astype(str).str.lower()
            if all(field in row.tolist() for field in self.required_fiels):
                return i
        return None

    def import_students(self, file):
        rapport = []
        required_fields = ['matricule', 'noms', 'date de naissance', 'lieu de naissance', 'sexe']
        try:
            df = read_excel(file, header=None, engine="openpyxl")
            header_row = self.detect_header_row(df)
            if header_row is None:
                rapport.append([False, f"Entêtes introuvables : {self.required_fiels}"])
            else:
                from osm.utils import one_escape, is_alphanumeric
                from pandas import isnull, to_datetime, NaT, notna
                df.columns = df.iloc[header_row].fillna('').astype(str).str.lower()
                df = df[(header_row + 1):].reset_index(drop=True)
                saved_lines = 0
                for index, line in df.iterrows():
                    line_number = header_row + 2 + index
                    if line.isnull()[required_fields].any():
                        empty_fields = line[required_fields].isnull()
                        empty_fields_list = empty_fields[empty_fields].index.tolist()
                        if len(empty_fields_list) == 1:
                            error = f"Valeur manquante d'un champ obligatoire : {empty_fields_list[0]}"
                        else:
                            error = f"Valeur manquante de {len(empty_fields_list)} champs obligatoires : " \
                                    f"{empty_fields_list}"
                        rapport.append([False, f"Ligne {line_number} : {error}"])
                    else:
                        matricule = line['matricule']
                        matricule_str = str(matricule)
                        if len(matricule_str) != 9 or not matricule_str.isnumeric():
                            rapport.append([False, f"Ligne {line_number} : Le matricule doit être une suite de 9 "
                                                   f"chiffres"])
                            continue
                        if Student.objects_all.filter(unique_id=matricule).exists():
                            rapport.append([False, f"Ligne {line_number} : Ce matricule a déjà été enregistré pour un "
                                                   f"autre élève"])
                            continue
                        nom = one_escape(str(line['noms'])).upper()
                        if not is_alphanumeric(nom):
                            rapport.append([False, f"Ligne {line_number} : Le nom doit être une chaîne alphanumérique"])
                            continue
                        prenom = one_escape(str(line['prénoms'])).title() if not isnull(line.get('prénoms')) else ''
                        if not is_alphanumeric(prenom):
                            rapport.append([False, f"Ligne {line_number} : Le prénom doit être une chaîne "
                                                   f"alphanumérique"])
                            continue
                        date_naissance = self.parse_birthdate(line['date de naissance'])
                        if date_naissance is None:
                            rapport.append([False, f"Ligne {line_number} : La date de naissance est invalide "
                                                   f"(format attendu : jj/mm/aaaa)"])
                            continue
                        now = datetime.now().year
                        min_year, max_year = now - 30, now - 8
                        if not (min_year <= date_naissance.year <= max_year):
                            rapport.append([False, f"Ligne {line_number} : L'année de naissance doit être comprise "
                                                   f"entre {min_year} et {max_year} (inclus)"])
                            continue
                        if Student.objects_all.filter(nom=nom, prenom=prenom, date_naissance=date_naissance).exists():
                            rapport.append([False, f"Ligne {line_number} : Un(e) élève du même nom et né le même jour "
                                                   f"a déjà été enregistré"])
                            continue
                        lieu_naissance = str(line['lieu de naissance']).title()
                        sexe = ''
                        if str(line['sexe']).lower() in ["fille", "féminin", "femme", "f"]:
                            sexe = "Fille"
                        elif str(line['sexe']).lower() in ["garçon", "masculin", "homme", "h", "g", 'm']:
                            sexe = "Garçon"
                        if not sexe:
                            rapport.append([False, f"Ligne {line_number} : La valeur du sexe est incorrecte"])
                            continue
                        classe_id = None
                        if 'classe' in df.columns and notna(line['classe']):
                            raw_classe = str(line['classe']).strip()
                            if raw_classe and raw_classe != "(Aucune)":
                                classroom = ClassRoom.objects.filter(code__iexact=raw_classe).first()
                                if classroom is None:
                                    rapport.append([False, f"Ligne {line_number} : classe « {raw_classe} » "
                                                           f"inconnue — élève importé sans classe (à régulariser)."])
                                else:
                                    classe_id = classroom.id
                        statut = "Nouveau"
                        if 'statut' in df.columns and notna(line['statut']):
                            if str(line['statut']).lower() in ["redoublant", "redoublante", "r"]:
                                statut = "Redoublant"
                        Student.objects.create(nom=nom, prenom=prenom, statut=statut, sexe=sexe, classe_id=classe_id,
                                               date_naissance=date_naissance, lieu_naissance=lieu_naissance,
                                               unique_id=matricule)
                        saved_lines += 1
                if saved_lines:
                    rapport.insert(0, [True,f"{'Une ligne insérée' if saved_lines == 1 else f'{saved_lines} lignes insérées'} avec succès"])
                elif not rapport:
                    rapport.append([False, "Il semble que ce fichier ne contienne aucune donnée d'élève"])
        except:
            rapport.append((False, "Fichier illisible, vérifier qu'il soit bien au format xls, xlsx ou odt"))
        return rapport


# Affichage de la liste des élèves
class Students(ListView):
    model = Student
    template_name = "students_list.html"
    title = "Liste des Élèves"
    objects = "élève(s)"


class Parents(ListView):
    template_name = "parents_list.html"
    title = "Liste des Parents d'Élèves"
    model = Parent
    objects = "parents d'élève(s)"


class StudentAdd(LoggedAdminView):
    template_name = "add_student.html"
    title = "Ajout d'un Élève"

    def get(self, *args, **kwargs):
        student_form = StudentForm(context={'request': self.request})
        context = {"title": self.title, "form": student_form}
        return render(self.request, self.template_name, context)

    def post(self, *args, **kwargs):
        student_form = StudentForm(self.request.POST, self.request.FILES, context={'request': self.request})
        context = {"title": self.title, "form": student_form}
        if student_form.is_valid():
            student_form.save()
            message(self.request, "Élève ajouté avec succès.")
            return redirect("students")
        return render(self.request, self.template_name, context)


class ParentAdd(LoggedAdminView):
    template_name = "add_parent.html"
    title = "Ajout d'un Parent d'Élève"

    def get(self, *args, **kwargs):
        parent_form = ParentForm(context={'request': self.request})
        context = {"title": self.title, "form": parent_form, "reset": "Tout effacer"}
        return render(self.request, self.template_name, context)

    def post(self, *args, **kwargs):
        parent_form = ParentForm(self.request.POST, context={'request': self.request})
        context = {"title": self.title, "form": parent_form, "reset": "Tout effacer"}
        if parent_form.is_valid():
            parent_form.save()
            message(self.request, "Parent ajouté avec succès.")
            return redirect("parents")
        return render(self.request, self.template_name, context)


class ParentEdit(LoggedAdminView):
    template_name = "add_parent.html"
    title = "Modification des informations"

    def get(self, *args, **kwargs):
        form = ParentForm(context={'request': self.request}, instance=self.get_object())
        return render(self.request, self.template_name, {"title": self.title, "form": form,
                                                         "reset": "Annuler les changements"})

    def get_object(self):
        parent_id = self.kwargs.get("id")
        parents = Parent.objects
        return get_object_or_404(parents, pk=parent_id)

    def post(self, *args, **kwargs):
        instance = self.get_object()
        default = model_to_dict(instance)
        form = ParentForm(self.request.POST, context={'request': self.request}, instance=instance)
        if form.is_valid():
            parent = form.save()
            parent = model_to_dict(parent)
            if parent != default:
                message(self.request, "Informations du parent modifiées avec succès !")
            return redirect("parent-details", id=instance.pk)
        return render(self.request, self.template_name, {"form": form, "title": self.title,
                                                         "reset": "Annuler les changements"})


class StudentEdit(LoggedAdminView):
    template_name = "add_student.html"
    title = "Modification des informations"
    nb = "NB : Si vous modifiez la salle de classe d'un élève, toutes ses notes préalablement enregistrées " \
         "seront perdues."

    def get(self, *args, **kwargs):
        form = StudentForm(context={'request': self.request}, instance=self.get_object())
        return render(self.request, self.template_name, {"title": self.title, "form": form, "reset": "Par défaut",
                                                         "nb": self.nb})

    def get_object(self):
        student_id = self.kwargs.get("id")
        students = Student.objects_all.select_related('classe', 'pere', 'mere')
        return get_object_or_404(students, pk=student_id)

    def post(self, *args, **kwargs):
        from osm.utils import delete_image
        default = self.get_object()
        default_pk = default.pk
        old_image = default.photo
        form = StudentForm(self.request.POST, self.request.FILES, context={'request': self.request},
                           instance=default)
        default_classroom = default.classe.classe.__str__() if default.classe else None
        default = model_to_dict(default)
        if form.is_valid():
            student = form.save()
            image = form.cleaned_data["photo"]
            if old_image and old_image != image:
                delete_image(old_image)
            """
            TODO
            """
            if student.classe and (student.classe.classe.__str__() != default_classroom):
                notes = Note.objects.filter(eleve=student)
                notes.delete()
            student = model_to_dict(student)
            if student != default:
                message(self.request, "Élève modifié avec succès.")
            else:
                message(self.request, "Aucune modification effectuée.", msg_type="warning")
            return redirect("student-details", id=default_pk)
        return render(self.request, self.template_name, {"form": form, "title": self.title, "reset": "Par défaut",
                                                         "nb": self.nb})


class ParentDelete(DeleteView):
    success_url = "parents"
    alerte = "des parents d'élèves ?"
    model = Parent
    title = "Suppression d'un Parent d'Élève"
    message = "Parent d'Élève supprimé avec succès."


class StudentDelete(DeleteView):
    success_url = "students"
    alerte = "des élèves ?"
    model = Student
    title = "Suppression d'un Élève"
    message = "Élève supprimé avec succès."


class StudentDetails(ADetailView):
    title = "Détails sur l'Élève"
    template_name = "student_details.html"
    model = Student


class ParentDetails(ADetailView):
    title = "Détails sur le Parent d'Élève"
    template_name = "parent_details.html"
    model = Parent


@with_users_school_schema
def discipline(request):
    def student_list(search):
        students_set = list()
        students_list = Student.objects.filter(Q(unique_id__icontains=search) | Q(nom__icontains=search) |
                                               Q(prenom__icontains=search) | Q(classe__in=ClassRoom.objects.filter(code__icontains=search)),
                                               classe__isnull=False).order_by_classroom_level()
        for student in students_list:
            students_set.append(student.dstd)
        return students_set

    students, info, trim = [], "", 0
    search_form = SearchForm(context={'trim': True})
    if request.method == "POST":
        search_form = SearchForm(request.POST or None, context={'trim': True})
        if search_form.is_valid() and search_form.cleaned_data['search']:
            students = student_list(search=search_form.cleaned_data['search'])
            trim = search_form.cleaned_data['trimestre']
            info = (f"{len(students)} élève(s) trouvé(s).", "Aucun(e) élève trouvé")[len(students) == 0]
    context = {'students': students, 'info': info, 'title': "Gestion de la discipline", 'trim': trim,
               'form': search_form}
    return render(request, "discipline.html", context)


class Discipline(LoggedAdminView):
    template_name = "discipline_edit.html"

    # TODO check period
    def check_period(self):
        pass

    def get(self, *args, **kwargs):
        sid, trim = self.kwargs.get('id'), self.kwargs.get('trim')
        student = Student.objects.get(pk=sid)
        std = student.std(trim)
        std_info = f"Données disciplinaires pour le trimestre {trim} de l'élève {student.dstd['nom']}"
        if std:
            dform = DForm(instance=std)
        else:
            dform = DForm()
        context = {'form': dform, 'sid': sid, 'trim': trim, 'std_info': std_info, 'show': True}
        return render(self.request, self.template_name, context)

    def post(self, *args, **kwargs):
        sid, trim = self.kwargs.get('id'), self.kwargs.get('trim')
        student = Student.objects.get(pk=sid)
        std = student.std(trim)
        std_info = f"Données disciplinaires pour le trimestre {trim} de l'élève {student.dstd['nom']}"
        if std:
            dform = DForm(self.request.POST, instance=std, context={'request': self.request})
        else:
            dform = DForm(self.request.POST, context={'request': self.request})
        if dform.is_valid() and dform.has_changed():
            avert, blame, excl = False, False, 0
            cons, absnj = dform.cleaned_data['cons'], dform.cleaned_data['abs'] - dform.cleaned_data['absj']
            if (6 <= absnj <= 9) or (6 <= cons <= 9):
                avert = True
            elif (10 <= absnj <= 14) or (10 <= cons <= 14):
                blame = True
            elif (15 <= absnj <= 18) or (15 <= cons <= 18):
                excl = 3
            elif (19 <= absnj <= 25) or (19 <= cons <= 25):
                excl = 5
            elif (25 <= absnj <= 29) or (25 <= cons <= 29):
                excl = 8
            if std:
                std.avert, std.retards = avert, dform.cleaned_data['retards']
                std.excl, std.excl_def = excl, dform.cleaned_data['excl_def']
                std.blame, std.cons = blame, dform.cleaned_data['cons']
                std.abs, std.absj = dform.cleaned_data['abs'], dform.cleaned_data['absj']
                std.save()
            else:
                dstd = StudentDiscipline(trim=trim, student=student, avert=avert,
                                         excl_def=dform.cleaned_data['excl_def'],
                                         excl=excl, blame=blame, cons=dform.cleaned_data['cons'],
                                         abs=dform.cleaned_data['abs'], absj=dform.cleaned_data['absj'],
                                         retards=dform.cleaned_data['retards'])
                dstd.save()
            message(self.request, "Données enregistrées avec succès.")
            student.classe.touch_notes(term_index=trim)
        else:
            message(self.request, "Aucune donnée modifiée ou enregistrée.", msg_type="warning")
        context = {'form': dform, 'sid': sid, 'trim': trim, 'std_info': std_info, 'show': False}
        response = render(self.request, self.template_name, context)
        response['HX-Trigger'] = 'AJAXMessages'
        return response


"""
=============================================================================
 STATISTIQUES — Répartition des élèves par ÂGE et par SEXE
=============================================================================
 Tableau croisé : lignes = classes (+ Total), colonnes = âges présents
 (chacun éclaté F/G/T) + bloc Total (F/G/T).
 Âge = âge au 31 DÉCEMBRE de l'année scolaire (convention officielle).
 Vue web + export PDF + export Excel.
=============================================================================
"""


def _ref_year_dec31(school_year_str):
    """Année scolaire '2025/2026' -> 2025 (le 31/12 de l'année de RENTRÉE)."""
    try:
        return int(school_year_str.split("/")[0])
    except Exception:
        return date.today().year


def age_at(dob, ref_year):
    """Âge révolu au 31 décembre de ref_year."""
    if not dob:
        return None
    ref = date(ref_year, 12, 31)
    return ref.year - dob.year - ((ref.month, ref.day) < (dob.month, dob.day))


def build_age_sex_table(year=None):
    """Construit la structure du tableau croisé âge×sexe par classe.
    Renvoie un dict prêt pour le template ET les exports."""
    year = year or school_year()
    ref = _ref_year_dec31(year)

    classrooms = list(ClassRoom.objects.all().order_by_niveau())

    # 1er passage : collecter les âges présents + compter par (classe, âge, sexe)
    ages = set()
    # counts[classroom_id][age] = {"F": n, "G": n}
    counts = {c.id: {} for c in classrooms}
    for c in classrooms:
        for st in c.students.all():
            a = age_at(st.date_naissance, ref)
            if a is None:
                continue
            ages.add(a)
            slot = counts[c.id].setdefault(a, {"F": 0, "G": 0})
            if st.sexe == "Fille":
                slot["F"] += 1
            else:
                slot["G"] += 1

    ages = sorted(ages)

    # 2e passage : lignes par classe avec F/G/T par âge + total ligne
    rows = []
    col_tot = {a: {"F": 0, "G": 0} for a in ages}   # totaux par âge (bas)
    grand = {"F": 0, "G": 0}
    for c in classrooms:
        cells = []
        rF = rG = 0
        for a in ages:
            slot = counts[c.id].get(a, {"F": 0, "G": 0})
            f, g = slot["F"], slot["G"]
            cells.append({"F": f, "G": g, "T": f + g})
            rF += f; rG += g
            col_tot[a]["F"] += f; col_tot[a]["G"] += g
        rows.append({"classe": c.code, "cells": cells, "F": rF, "G": rG, "T": rF + rG})
        grand["F"] += rF; grand["G"] += rG

    # ligne des totaux (bas)
    total_cells = [{"F": col_tot[a]["F"], "G": col_tot[a]["G"], "T": col_tot[a]["F"] + col_tot[a]["G"]} for a in ages]
    total_row = {"classe": "TOTAL", "cells": total_cells, "F": grand["F"], "G": grand["G"], "T": grand["F"] + grand["G"]}

    return {"year": year, "ref_year": ref, "ages": ages, "rows": rows, "total_row": total_row}


# ---------------------------------------------------------------------------
#  VUE STATISTIQUES — Répartition des élèves par ÂGE et par SEXE
# ---------------------------------------------------------------------------
@logged_admin_view
def age_sex_stats(request):
    data = build_age_sex_table(school_year())
    data['title'] = "Répartition par âge et par sexe"
    return render(request, "age_sex.html", data)


# ---------------------------------------------------------------------------
#  EXPORT PDF STATISTIQUES — Répartition des élèves par ÂGE et par SEXE
# ---------------------------------------------------------------------------
@logged_admin_view
def age_sex_stats_pdf(request):
    data = build_age_sex_table(school_year())
    if not Student.objects.exists() or not ClassRoom.objects.exists():
        if not Student.objects.exists():
            msg = "Aucun élève enregistré."
        else:
            msg = "Aucune classe enregistrée."
        message(request, msg, msg_type="error")
        return safe_redirect_back(request)
    pdf = AgeSexTablePDF(data, request.user.school)
    return pdf_response(pdf, f"Répartition par Age et par Sexe {data['year']}.pdf")


# ---------------------------------------------------------------------------
#  EXPORT EXCEL STATISTIQUES — Répartition des élèves par ÂGE et par SEXE
# ---------------------------------------------------------------------------
@logged_admin_view
def age_sex_stats_xlsx(request):
    data = build_age_sex_table(school_year())
    ages, rows, total = data["ages"], data["rows"], data["total_row"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Répartition âge-sexe"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    fill = PatternFill("solid", fgColor="1B3A57")
    white_bold = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Ligne 1 : Classe | <age> (fusion 3) ... | Total (fusion 3)
    ws.cell(1, 1, "Classe"); ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    col = 2
    for a in ages:
        ws.cell(1, col, f"{a} ans")
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        col += 3
    ws.cell(1, col, "Total")
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)

    # Ligne 2 : F G T répétés
    col = 2
    for _ in range(len(ages) + 1):     # +1 pour le bloc Total
        ws.cell(2, col, "F"); ws.cell(2, col + 1, "G"); ws.cell(2, col + 2, "T")
        col += 3

    # style en-têtes
    for r in (1, 2):
        for cc in range(1, col):
            cell = ws.cell(r, cc)
            cell.font = white_bold; cell.alignment = center
            cell.fill = fill; cell.border = border

    # lignes de données
    rownum = 3
    for row in rows + [total]:
        ws.cell(rownum, 1, row["classe"])
        if row["classe"] == "TOTAL":
            ws.cell(rownum, 1).font = bold
        cc = 2
        for cell in row["cells"]:
            ws.cell(rownum, cc, cell["F"] or "")
            ws.cell(rownum, cc + 1, cell["G"] or "")
            ws.cell(rownum, cc + 2, cell["T"] or "")
            cc += 3
        ws.cell(rownum, cc, row["F"]); ws.cell(rownum, cc + 1, row["G"])
        ws.cell(rownum, cc + 2, row["T"])
        for c2 in range(1, cc + 3):
            cell = ws.cell(rownum, c2)
            cell.alignment = center; cell.border = border
            if row["classe"] == "TOTAL":
                cell.font = bold
        rownum += 1

    ws.column_dimensions["A"].width = 16
    for i in range(2, col + 3):
        ws.column_dimensions[get_column_letter(i)].width = 5

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="Répartition par Age et par Sexe {data["year"]}.xlsx"'
    wb.save(resp)
    return resp


"""
=============================================================================
 CERTIFICAT DE SCOLARITÉ
=============================================================================
 Conditionné à la solvabilité : l'élève ne doit avoir AUCUN reste sur les
 frais qui ENTRENT EN CAISSE (affects_cashbox=True).
=============================================================================
"""


# ---------------------------------------------------------------------------
#  HELPER SOLVABILITÉ
# ---------------------------------------------------------------------------
def unpaid_cashbox_fees(student, year):
    """Liste des frais EN CAISSE non soldés de l'élève pour l'année.
    Vide => l'élève est à jour (éligible au certificat)."""
    try:
        return [r for r in student.student_fee_status(year) if r["affects_cashbox"] and r["reste"] > 0]
    except Exception:
        return []


def is_solvent(student, year):
    return not unpaid_cashbox_fees(student, year)


def _available_years():
    """Années présentes en base + année courante, la plus récente d'abord."""
    years = list(StudentEnrollment.objects.values_list("school_year__libelle", flat=True).distinct())
    current = school_year()
    if current not in years:
        years.append(current)
    return sorted({y for y in years if y}, reverse=True)


def _classroom_for(student, year):
    """Classe de l'élève POUR L'ANNÉE DEMANDÉE (via son inscription)."""
    enr = (StudentEnrollment.objects.filter(student=student, school_year__libelle=year).select_related("classroom").first())
    if enr and enr.classroom_id:
        return enr.classroom
    # repli : année courante -> classe actuelle
    return student.classe if year == school_year() else None


# ---------------------------------------------------------------------------
#  VUE CERTIFICAT DE SCOLARITÉ : ON CHOISIT L'ÉLÈVE
# ---------------------------------------------------------------------------
@logged_admin_view
def certificates(request):
    years = _available_years()
    year = request.GET.get("year") or school_year()
    if year not in years:
        year = years[0] if years else school_year()

    q = (request.GET.get("q") or "").strip()
    rows = []
    if q:
        students = (Student.objects.filter(Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(unique_id__icontains=q))
                    .select_related("classe").order_by("nom", "prenom")[:30])
        for st in students:
            classroom = _classroom_for(st, year)
            unpaid = unpaid_cashbox_fees(st, year)
            rows.append({
                "student": st,
                "classroom": classroom,
                "enrolled": classroom is not None,
                "solvent": not unpaid,
                "reste": sum(r["reste"] for r in unpaid),
            })

    return render(request, "certificates.html", {
        "years": years, "year": year, "q": q, "rows": rows, 'title': "Certificat de Scolarité"
    })


# ---------------------------------------------------------------------------
#  TÉLÉCHARGEMENT CERTIFICAT DE SCOLARITÉ : ANNÉE PASSÉE EN PARAMÈTRE
# ---------------------------------------------------------------------------
@logged_admin_view
def enrollment_certificates(request, student_id):
    """Certificat nominatif pour l'année passée en paramètre (?year=)."""
    student = get_object_or_404(Student, pk=student_id)
    year = request.GET.get("year") or school_year()

    unpaid = unpaid_cashbox_fees(student, year)
    if unpaid:
        reste = sum(r["reste"] for r in unpaid)
        details = ", ".join(f"{r['fee_type']} ({r['reste']:,} F)".replace(",", " ")
                            for r in unpaid)
        message(request, f"Certificat indisponible : {student.short_name} n'est pas à jour pour "
                f"{year} (reste {reste:,} FCFA — {details}).".replace(",", " "), msg_type="error")
        return safe_redirect_back(request, "certificates")
    classroom = _classroom_for(student, year)
    if classroom is None:
        message(request, f"{student.short_name} n'a pas d'inscription enregistrée pour {year}.", msg_type="error")
        return safe_redirect_back(request, "certificates")

    pdf = EnrollmentCertificate(student, year, request.user.school, classroom=classroom)
    return pdf_response(pdf, f"Certificat de Scolarite {student.short_name} {year}.pdf")


@logged_admin_view
def enrollment_certificate_blank(request):
    """Certificat VIERGE : toutes les mentions variables en lignes à remplir.
    Aucune condition de solvabilité (aucun élève n'est désigné)."""
    pdf = EnrollmentCertificate(None, None, request.user.school, blank=True)
    return pdf_response(pdf, f"Certificat de Scolarité vierge.pdf")


# ---------------------------------------------------------------------------
#  VUE CERTIFICAT DE SCOLARITÉ POUR UN ÉLÈVE POUR L'ANNÉE COURANTE
# ---------------------------------------------------------------------------
@logged_admin_view
def enrollment_certificate(request, id):
    student = get_object_or_404(Student.objects.select_related('classe'), pk=id)
    year = school_year()
    unpaid = unpaid_cashbox_fees(student, year)
    if unpaid:
        reste = sum(r["reste"] for r in unpaid)
        details = ", ".join(f"{r['fee_type']} ({r['reste']:,} F)".replace(",", " ") for r in unpaid)
        message(request, f"Certificat indisponible : {student.short_name} n'est pas à jour "
                f"(reste {reste:,} FCFA — {details}).".replace(",", " "), msg_type="error")
        return safe_redirect_back(request)

    pdf = EnrollmentCertificate(student, year, request.user.school, student.classe)
    return pdf_response(pdf, f"Certificat de Scolarité {student.short_name}.pdf")


# Couleurs
GREEN = (10, 125, 63)
HEAD  = (27, 58, 87)
GREY  = (110, 120, 132)
LINE  = (170, 180, 190)
RED = (210, 31, 60)
YELLOW = (249, 214, 22)
INK = (21, 35, 59)
BLUE = (10, 61, 98)
NAVY  = (10, 61, 98)
DARK  = (30, 40, 55)


"""
=============================================================================
 EnrollmentCertificate — Certificat de scolarité
=============================================================================
 1. ANNÉE CIBLE : le certificat peut porter sur une année PASSÉE. La classe
    n'est alors pas forcément student.classe (classe actuelle) mais peut être
    celle de l'inscription de l'année demandée -> paramètre `classroom`.

 2. VERSION VIERGE (`blank=True`) : toutes les valeurs sont remplacées par
    des LIGNES à remplir à la main, y compris le cadre photo. L'établissement
    peut ainsi imprimer une réserve de certificats. `student` vaut None dans
    ce mode.
=============================================================================
"""
class EnrollmentCertificate(FPDF):
    """pdf = EnrollmentCertificate(student, year, school) -> pdf_response(...)
    Version vierge : EnrollmentCertificate(None, year, school, blank=True)"""

    L, R = 10, 200
    GREY = (120, 130, 142)
    LINE = (200, 208, 216)

    def __init__(self, student, year, school, classroom=None, blank=False):
        super().__init__(orientation="P", unit="mm", format="A4")
        add_fonts(self)
        self.set_auto_page_break(False)
        self.set_margins(6, 16, 6)
        self.student = student
        self.year = year
        self.school = school
        self.blank = blank
        # classe de l'ANNÉE DEMANDÉE (peut différer de la classe actuelle)
        self.classroom = classroom or (getattr(student, "classe", None) if student else None)
        try:
            self._cachet_bytes = stamp_bytes(school.cachet)
            self._visa_bytes = stamp_bytes(school.visa)
        except Exception:
            self._cachet_bytes = None
            self._visa_bytes = None
        self.add_page()
        filigrane(self, x=50, y=95, w=110)
        self.set_font("inter", "", 8)
        base_header(self, mode="P", y_img=10)
        self._title()
        self._body()

    # ------------------------------------------------------------------
    #  HELPERS DE MISE EN FORME
    # ------------------------------------------------------------------
    def _label(self, x, y, fr, en):
        """Libellé bilingue : français puis anglais en italique dessous.
        Renvoie la largeur occupée (pour poser la valeur juste après)."""
        self.set_font("inter", "", 10)
        self.set_text_color(*DARK)
        w_fr = self.get_string_width(fr)
        self.set_xy(x, y)
        self.cell(w_fr + 1, 4.6, fr)
        self.set_font("inter", "I", 8)
        self.set_text_color(*self.GREY)
        w_en = self.get_string_width(en)
        self.set_xy(x, y + 4.3)
        self.cell(w_en + 1, 3.4, en)
        return max(w_fr, w_en) + 3

    def _rule(self, x, y, w):
        """Ligne à remplir à la main (version vierge)."""
        self.set_draw_color(*self.GREY)
        self.set_line_width(0.25)
        self.line(x, y + 4.4, x + max(10, w) if w else 198, y + 4.4)

    def _value(self, x, y, text, size=10.5, color=NAVY, line_w=None):
        """Valeur en gras — ou LIGNE À REMPLIR si version vierge.
        `line_w` : longueur du trait ; par défaut jusqu'à la marge droite."""
        if self.blank:
            self._rule(x, y, line_w)
            return
        self.set_font("inter", "B", size)
        self.set_text_color(*color)
        self.set_xy(x, y)
        self.cell(0, 4.6, str(text) if text not in (None, "") else "")

    def _field(self, x, y, fr, en, value, size=10.5, line_w=None):
        """Libellé bilingue + valeur (ou ligne). Renvoie la largeur du libellé."""
        w = self._label(x, y, fr, en)
        self._value(x + w, y, value, size, line_w=line_w)
        return w

    # ------------------------------------------------------------------
    def _title(self):
        y = 60
        self.set_font("inter", "B", 17)
        self.set_text_color(*GREEN)
        self.set_xy(self.L, y)
        self.cell(self.R - self.L, 8, "CERTIFICAT DE SCOLARITÉ", align="C")
        self.set_font("inter", "I", 10)
        self.set_text_color(*self.GREY)
        self.set_xy(self.L, y + 8)
        self.cell(self.R - self.L, 5, "SCHOOL ATTENDANCE CERTIFICATE", align="C")
        # filet tricolore
        cx = (self.L + self.R) / 2
        self.set_line_width(1)
        self.set_draw_color(*GREEN)
        self.line(cx - 35, y + 15, cx - 35 + 70 / 3, y + 15)
        self.set_draw_color(*RED)
        self.line(cx - 35 + 70 / 3, y + 15, cx - 35 + 140 / 3, y + 15)
        self.set_draw_color(*YELLOW)
        self.line(cx - 35 + 140 / 3, y + 15, cx + 35, y + 15)

    # ------------------------------------------------------------------
    #  CORPS
    # ------------------------------------------------------------------
    def _body(self):
        from staff.models import Personnel
        st = self.student
        s = self.school
        L, R = self.L, self.R

        # --- photo (cadre vide et légendé en version vierge) -----------
        ph_w, ph_h = 30, 40
        ph_x, ph_y = R - ph_w - 5, 85
        if not self.blank:
            try:
                src = (st.photo if getattr(st, "photo", None) else "static/image/student.jpg")
                photo = resize_image(src, id_card=True, ratio=(30, 40))
                self.image(photo, x=ph_x, y=ph_y, w=ph_w, h=ph_h)
            except Exception:
                pass
        else:
            self.set_font("inter", "I", 7)
            self.set_text_color(*self.GREY)
            self.set_xy(ph_x, ph_y + ph_h / 2 - 3)
            self.cell(ph_w, 5, "Photo", align="C")
        self.set_draw_color(*self.LINE)
        self.set_line_width(0.3)
        self.rect(ph_x, ph_y, ph_w, ph_h)

        text_r = ph_x - 6

        # --- référence de registre -------------------------------------
        self.set_font("inter", "", 8)
        self.set_text_color(*self.GREY)
        self.set_xy(L, 85)
        self.cell(60, 4, "Réf. N° ___________________________________")

        # --- 1. le signataire ------------------------------------------
        y = 95
        chef = Personnel.objects.filter(poste="Chef d'Établissement").first()
        add = ("e" if chef.civilite == "Madame" else "") if (not self.blank and chef) else "(e)"
        w = self._label(L, y, f"Je soussigné{add},", "I the undersigned,")
        self._value(L + w, y, chef.__str__().upper() if chef else "", line_w=text_r - (L + w))

        y += 12
        poste = s.chef
        w = self._label(L, y, f"{poste} du", "The Principal of")
        self.set_font("inter", "B", 10)
        self.set_text_color(*NAVY)
        self.set_xy(L + w, y)
        self.multi_cell(text_r - (L + w), 4.6, (s.nom or "").upper())
        self.set_xy(L + w, y + 4.6)
        self.set_font("inter", "I", 8)
        self.set_text_color(*self.GREY)
        self.cell(0, 3.4, (s.name or "").upper())

        # --- 2. l'élève -------------------------------------------------
        y += 12
        w = self._label(L, y, "Attestons que l'élève", "Certify that the student")
        if self.blank:
            self._rule(L + w, y, text_r - (L + w))
            y += 11
        else:
            self.set_font("inter", "B", 11)
            self.set_text_color(*NAVY)
            self.set_xy(L + w, y)
            self.multi_cell(text_r - (L + w), 4.8, str(st).upper())
            y = max(self.get_y() + 4, y + 11)

        # naissance
        try:
            date_naissance = format_date(st.date_naissance)
        except Exception:
            date_naissance = ""
        plus = ("e" if st.sexe == "Fille" else "") if not self.blank else "(e)"
        w = self._field(L, y, f"Né{plus} le", "Born on", date_naissance, line_w=42)
        x2 = L + 78
        self._field(x2, y, "à", "at", getattr(st, "lieu_naissance", "") if st else "", line_w=text_r - x2 - 12)

        # --- 3. l'inscription -------------------------------------------
        y += 12
        self._label(L, y, f"Est régulièrement inscrit{plus} comme élève au sein de notre établissement",
                    "Is duly enrolled as a student at our school")

        y += 12
        w = self._label(L, y, "En classe de", "In the class of")
        classe = self.classroom.code if self.classroom else ""
        self._value(L + w, y, classe, color=GREEN, size=11, line_w=60)

        # --- 4. année + matricule ---------------------------------------
        y += 13
        w = self._field(L, y, "Année scolaire", "Academic year", self.year, line_w=34)
        self._field(L + 78, y, "sous le matricule numéro", "Registration number",
                    getattr(st, "unique_id", "") if st else "", size=10, line_w=40)

        # --- 5. formule de délivrance -----------------------------------
        y += 14
        self.set_font("inter", "", 10)
        self.set_text_color(*DARK)
        self.set_xy(L, y)
        self.multi_cell(R - L, 4.6, "En foi de quoi le présent certificat lui est délivré pour servir et valoir "
                                    "ce que de droit.", align="L")
        self.set_font("inter", "I", 8)
        self.set_text_color(*self.GREY)
        self.set_xy(L, self.get_y() + 0.5)
        self.multi_cell(R - L, 4.3, "In witness whereof the present certificate is issued to serve and avail as "
                                    "of right.", align="L")

        # --- 6. lieu, date, signature, cachet ---------------------------
        y = self.get_y() + 14
        localite = s.localite
        w = self._label(R - 82, y, f"Fait à {localite}, le", f"Done in {localite}, on")
        # date en rouge seulement si les tampons sont là ET hors version vierge
        if self._visa_bytes and self._cachet_bytes and not self.blank:
            self._value(R - 82 + w, y, f"{date.today():%d/%m/%Y}", size=10, color=RED)
        else:
            self._rule(R - 82 + w, y, 30)

        y += 13
        self.set_font("inter", "B", 10)
        self.set_text_color(*NAVY)
        self.set_xy(R - 82, y)
        self.cell(82, 5, f"Le {poste}", align="C")
        self.set_font("inter", "I", 7)
        self.set_text_color(*self.GREY)
        self.set_xy(R - 82, y + 4.6)
        self.cell(82, 4, "The Principal", align="C")

        # cachet + visa : jamais sur une version vierge
        if not self.blank:
            paste_stamp(self, self._cachet_bytes, x=120, y=y + 10, w=40)
            paste_stamp(self, self._visa_bytes, x=155, y=y + 25, w=50)


"""
=============================================================================
 AgeSexTablePDF — Répartition des élèves par âge et par sexe (PDF)
=============================================================================
 A4 PAYSAGE. Double en-tête : ligne 1 = âges (+ Total), ligne 2 = F/G/T.
 Lignes = classes (+ ligne TOTAL).
=============================================================================
"""

class AgeSexTablePDF(FPDF):
    def __init__(self, data, school):
        super().__init__(orientation="L", unit="mm", format="A4")
        self.alias_nb_pages()
        add_fonts(self)
        self.set_margins(6, 6, 6)
        self.set_auto_page_break(True, margin=6)
        self.data = data
        self.now = datetime.now().strftime("%d-%m-%Y à %H:%M")
        self.school = school
        self.add_page()
        self.set_font("inter", "", 8)
        base_header(self, mode="L")
        self._title()
        self._table()

    def _title(self):
        self.ln(2)
        self.set_font("inter", "B", 12)
        self.set_text_color(*GREEN)
        self.cell(0, 7, "RÉPARTITION DES ÉLÈVES PAR ÂGE ET PAR SEXE", align="C")
        self.ln(6)
        self.set_font("inter", "", 8)
        self.set_text_color(*GREY)
        self.cell(0, 4, f"Année scolaire {self.data['year']} • Âge au 31 décembre {self.data['ref_year']} • "
                        f"F : Filles · G : Garçons · T : Total", align="C")
        self.ln(6)
        self.set_text_color(0)

    def _table(self):
        ages = self.data["ages"]
        rows = self.data["rows"] + [self.data["total_row"]]

        # largeurs : colonne classe large, puis 3 sous-colonnes par âge + Total
        page_w = 297 - 12
        class_w = 25
        n_blocks = len(ages) + 1                 # âges + bloc Total
        sub_w = max(6.0, (page_w - class_w) / (n_blocks * 3))
        col_widths = [class_w, ]
        for _ in range(n_blocks * 3):
            col_widths.append(sub_w)
        col_widths = tuple(col_widths)
        self.set_font("inter", "", 8)

        table = Table(self, line_height=5, col_widths=col_widths, text_align="CENTER", markdown=True,
                      repeat_headings=TableHeadingsDisplay.ON_TOP_OF_EVERY_PAGE, num_heading_rows=2)
        labels = [str(a) + " ans" for a in ages] + ["Total"]
        th1 = table.row()
        self.set_fill_color(*HEAD)
        self.set_text_color(255, 255, 255)
        # --- EN-TÊTE ligne 1 : Classe (fusion vert.) + âges + Total ---
        th1.cell("**Classe**", align="C", rowspan=2)
        for label in labels:
            th1.cell(f"**{label}**", align="C", colspan=3)

        # --- EN-TÊTE ligne 2 : F / G / T ---
        th2 = table.row()
        for _ in labels:
            th2.cell("**F**", align="C")
            th2.cell("**G**", align="C")
            th2.cell("**T**", align="C")

        # --- CORPS ---
        self.set_text_color(0)
        for row in rows:
            is_total = row["classe"] == "TOTAL"
            if is_total:
                self.set_font("inter", "B", 8)
                self.set_fill_color(230, 238, 233)
            else:
                self.set_font("inter", "", 8)
                self.set_fill_color(255, 255, 255)
            tr = table.row()
            # cellule classe
            tr.cell(row['classe'], align="L")
            # cellules F/G/T par âge + total
            for cell in row["cells"]:
                for key in ("F", "G", "T"):
                    value = cell[key]
                    tr.cell(str(value) if value else "", align="C")
            # bloc total de la ligne
            for key in ("F", "G", "T"):
                tr.cell(str(row[key]) if row[key] else "", align="C")
        table.render()

    def footer(self):
        self.set_y(-6)
        self.set_draw_color(200)
        self.line(6, 204, 291, 204)
        self.set_font("inter", "I", 7)
        self.set_text_color(*GREY)
        self.cell(145, 6, f"Document généré par Oméga School Manager le {self.now}", align="L")
        self.cell(140, 6, f"Répartition par âge et par sexe • Page {self.page_no()}/{{nb}}", align="R")


# Dimensions normalisées de la carte scolaire (norme ISO/IEC 7810 ID-1).
CARD_W, CARD_H = 85.6, 54.0

# Nouveau gris
GREY = (138, 147, 163)


class StudentsIdentityCardsCNI(FPDF):

    def __init__(self, *args, **kwargs):
        super().__init__(orientation='P', unit='mm', format='A4')
        self.add_font('inter', '', settings.INTER_REGULAR)
        self.add_font('inter', 'I', settings.INTER_ITALIC)
        self.add_font('inter', 'B', settings.INTER_BOLD)
        self.add_font('inter', 'BI', settings.INTER_BOLDITALIC)
        self.set_auto_page_break(auto=False)   # placement manuel précis
        self.set_margins(0, 0, 0)
        self.set_font('inter', '', 7)

        self.data = kwargs.pop('data')
        self.layout = self.data.get('layout', "sheet")

        # Logo établissement (flux NEUF à chaque génération -> pas de partage).
        self._logo = self._prepare_logo()
        # Cachet (rond établissement) ET visa (nominatif + signature) préchargés
        # EN BYTES une fois. On recrée un BytesIO neuf à chaque pose (jamais
        # épuisé, même en mode ZIP toutes les classes). Voir utils.stamp_bytes.
        self._cachet_bytes = stamp_bytes(self.data['school_data'].get('cachet'))  # noqa: F821
        self._visa_bytes = stamp_bytes(self.data['school_data'].get('visa'))  # noqa: F821

        students = self.data.get('students', [])
        self.default_student_photo = resize_image("static/image/student.jpg", id_card=True, ratio=(26, 30))
        if self.layout == 'single':
            self._render_single(students)
        else:
            self._render_sheet(students)

    # ------------------------------------------------------------------
    # Préparation du logo : on récupère un flux exploitable (ou None).
    # ------------------------------------------------------------------
    def _prepare_logo(self):
        logo = self.data['school_data'].get('logo')
        if not logo or logo == "static/image/no_image.jpg":
            return None
        try:
            # resize_image renvoie un BytesIO neuf -> sûr en génération en boucle.
            return resize_image(logo, new_width=300)
        except Exception:
            return None

    # ------------------------------------------------------------------
    # MODE PLANCHE A4 : grille 2 colonnes x 5 rangées (10 cartes / page).
    # Géométrie calculée pour des marges et gouttières régulières.
    # ------------------------------------------------------------------
    def _render_sheet(self, students):
        cols, rows = 2, 5
        per_page = cols * rows
        gutter_x, gutter_y = 6.0, 2.0
        margin_x = (self.w - cols * CARD_W - gutter_x) / 2 # 16.4 mm
        margin_y = (self.h - rows * CARD_H - gutter_y * (rows - 1)) / 2 # 9.5 mm

        for i, student in enumerate(students):
            pos = i % per_page
            if pos == 0:
                self.add_page()
            c = pos % cols
            r = pos // cols
            x = margin_x + c * (CARD_W + gutter_x)
            y = margin_y + r * (CARD_H + gutter_y)
            self._draw_card(student, x, y)
            self._cut_marks(x, y) # traits de découpe autour de la carte

    # ------------------------------------------------------------------
    # MODE 1 CARTE / PAGE : page au format exact ID-1 (imprimante à cartes).
    # ------------------------------------------------------------------
    def _render_single(self, students):
        for student in students:
            self.add_page(format=(CARD_W, CARD_H))
            self._draw_card(student, 0, 0)

    # ------------------------------------------------------------------
    # Traits de découpe : petits repères aux 4 coins de la carte (discrets,
    # pour guider le massicot sans tracer de cadre complet).
    # ------------------------------------------------------------------
    def _cut_marks(self, x, y, m=2.0):
        self.set_draw_color(150)
        self.set_line_width(0.1)
        # coin haut-gauche
        self.line(x - m, y, x, y); self.line(x, y - m, x, y)
        # haut-droit
        self.line(x + CARD_W, y, x + CARD_W + m, y); self.line(x + CARD_W, y - m, x + CARD_W, y)
        # bas-gauche
        self.line(x - m, y + CARD_H, x, y + CARD_H); self.line(x, y + CARD_H, x, y + CARD_H + m)
        # bas-droit
        self.line(x + CARD_W, y + CARD_H, x + CARD_W + m, y + CARD_H)
        self.line(x + CARD_W, y + CARD_H, x + CARD_W, y + CARD_H + m)

    # ==================================================================
    # DESSIN D'UNE CARTE à l'origine (x, y). Tout est en coordonnées
    # RELATIVES à (x, y) pour pouvoir la placer n'importe où sur la planche.
    # ==================================================================
    def _draw_card(self, student, x, y):
        sd = self.data['school_data']

        # --- Cadre léger de la carte (coins NON arrondis : arrondi à la découpe) ---
        self.set_draw_color(225)
        self.set_line_width(0.2)
        self.rect(x, y, CARD_W, CARD_H)

        # ---------- EN-TÊTE bilingue + logo central ----------
        self.set_xy(x + 2, y + 2.5)
        self.set_text_color(*BLUE)
        self.set_font('inter', 'B', 5.8)
        self.multi_cell(31, 2.6, "RÉPUBLIQUE DU CAMEROUN", align='C', new_x="RIGHT", new_y="TOP")
        self.set_xy(x + CARD_W - 33, y + 2.5)
        self.multi_cell(31, 2.6, "REPUBLIC OF CAMEROON", align='C', new_x="RIGHT", new_y="TOP")

        self.set_font('inter', '', 4.5)
        self.set_text_color(*GREY)
        self.set_xy(x + 2, y + 5.2)
        self.cell(31, 2, "Paix - Travail - Patrie", align='C')
        self.set_xy(x + CARD_W - 33, y + 5.2)
        self.cell(31, 2, "Peace - Work - Fatherland", align='C')

        # nom établissement (FR / EN) sous les devises
        self.set_font('inter', 'B', 5.5)
        self.set_text_color(*INK)
        self.set_xy(x + 2, y + 7.4)
        self.cell(31, 2.4, self._fit_shrink(sd.get('nom', ''), 31, 5.7, bold=True, min_size=4.8)[0], align='C')
        self.set_xy(x + CARD_W - 33, y + 7.4)
        self.cell(31, 2.4, self._fit_shrink(sd.get('name', ''), 31, 5.7, bold=True, min_size=4.8)[0], align='C')

        # logo central (si dispo)
        if self._logo:
            self.image(self._logo, x=x + CARD_W / 2 - 6, y=y + 2, w=12, h=12,
                       keep_aspect_ratio=True)

        # ---------- TITRE ----------
        #self.set_xy(x, y + 11.5)
        self.set_xy(x, y + 13.5)
        self.set_font('inter', 'B', 8)
        self.set_text_color(*GREEN)
        self.cell(CARD_W, 4, "CARTE D'IDENTITÉ SCOLAIRE / SCHOOl ID CARD", align='C')

        # ---------- PHOTO (≈ 1/3 de la largeur) ----------
        px, py = x + 3, y + 18.5
        pw, ph = 26, 30          # ~1/3 de 85.6 ; ratio identité
        photo = getattr(student, 'photo', None)
        self.image(resize_image(photo, id_card=True, ratio=(26, 30)) if photo else self.default_student_photo, x=px, y=py, w=pw, h=ph)

        # ---------- CHAMPS d'identité (à droite de la photo) ----------
        fx = px + pw + 3
        fw = x + CARD_W - fx - 3
        fy = py - 0.5

        def field(label_fr, label_en, value, fy_, value_color=INK, value_size=7.5):
            self.set_xy(fx, fy_)
            self.set_font('inter', '', 4.2)
            self.set_text_color(*GREY)
            self.cell(fw, 1.8, f"{label_fr}  /  {label_en}", align='L')
            txt, used_size = self._fit_shrink(value, fw, value_size, bold=True, min_size=6.5)
            self.set_xy(fx, fy_ + 1.9)
            self.set_font('inter', 'B', used_size)
            self.set_text_color(*value_color)
            self.cell(fw, 2.6, txt, align='L')
            return fy_ + 5.0  # hauteur d'un champ

        # Noms & Prénoms (sur une ligne, valeur en bleu)
        full_name = f"{student.nom} {student.prenom or ''}".strip()
        fy = field("Noms & Prénoms", "Name & First Names", full_name, fy, BLUE, 7.5)

        # Né(e) le + À (deux demi-colonnes)
        born = format_date(student.date_naissance)
        self.set_xy(fx, fy)
        self.set_font('inter', '', 4.2); self.set_text_color(*GREY)
        self.cell(fw / 2, 1.8, "Né(e) le  /  Born", align='L')
        self.cell(fw / 2, 1.8, "À  /  at", align='L')
        lieu_txt, lieu_size = self._fit_shrink(student.lieu_naissance or "—", fw / 2, 6.5, bold=True)
        self.set_xy(fx, fy + 1.9)
        self.set_font('inter', 'B', 6.5); self.set_text_color(*INK)
        self.cell(fw / 2, 2.4, born, align='L')
        self.set_font('inter', 'B', lieu_size)
        self.cell(fw / 2, 2.4, lieu_txt, align='L')
        fy += 5.0

        # Sexe + Classe
        self.set_xy(fx, fy)
        self.set_font('inter', '', 4.2); self.set_text_color(*GREY)
        self.cell(fw / 2, 1.8, "Sexe  /  Sex", align='L')
        self.cell(fw / 2, 1.8, "Classe  /  Class", align='L')
        self.set_xy(fx, fy + 1.9)
        self.set_font('inter', 'B', 6.5); self.set_text_color(*INK)
        self.cell(fw / 2, 2.4, str(getattr(student, 'sexe', '') or "—"), align='L')
        self.cell(fw / 2, 2.4, (student.classe.code if student.classe else "—"), align='L')
        fy += 5.0

        # Matricule
        self.set_xy(fx, fy)
        self.set_font('inter', '', 4.2); self.set_text_color(*GREY)
        self.cell(fw, 1.8, "Matricule  /  ID", align='L')
        self.set_xy(fx, fy + 1.9)
        self.set_font('inter', 'B', 7)
        #self.set_text_color(*RED)
        self.set_text_color(*BLUE)
        self.cell(fw, 2.6, str(student.unique_id or "—"), align='L')
        fy += 5.0

        # Contact d'urgence (en rouge)
        contacts_list = getattr(student, 'contacts_parent')
        contacts_list.append(sd.get('contact'))
        contacts = str(contacts_list[0]) if len(contacts_list) == 1 else f"{contacts_list[0]} / {contacts_list[1]}"
        self.set_xy(fx, fy)
        self.set_font('inter', '', 4.2); self.set_text_color(*GREY)
        self.cell(fw, 1.8, "Contact d'urgence  /  ICE", align='L')
        self.set_xy(fx, fy + 1.9)
        self.set_font('inter', 'B', 7)
        self.set_text_color(*RED)
        self.cell(fw, 2.6, contacts, align='L')

        # --- Cachet (rond établissement) + Visa (nominatif + signature) ---
        # Posés APRÈS les champs : peuvent chevaucher (authenticité). BytesIO
        # neuf à chaque pose via paste_stamp -> sûr en boucle.
        paste_stamp(self, self._cachet_bytes,  # noqa: F821
                    x=x + CARD_W - 38, y=y + 30, w=18)
        paste_stamp(self, self._visa_bytes,  # noqa: F821
                    x=x + CARD_W - 25, y=y + 35, w=22)
        """stamp_w = 19  # largeur discrète (mm)
        stamp_x = x + CARD_W - stamp_w - 3  # ancré à droite, marge 3mm
        stamp_y = y + 30"""

        # ---------- BANDE TRICOLORE + année (en bas) ----------
        self._footer_band(x, y)

        # reset couleur
        self.set_text_color(0)

    # ------------------------------------------------------------------
    # Bande tricolore en pied : vert (Année FR) | rouge+étoile | jaune (Year EN)
    # ------------------------------------------------------------------
    def _footer_band(self, x, y):
        band_h = 5.0
        by = y + CARD_H - band_h
        third = CARD_W / 3
        annee = self.data.get('annee', '')

        # vert
        self.set_fill_color(*GREEN)
        self.rect(x, by, third, band_h, style='F')
        # rouge (centre)
        self.set_fill_color(*RED)
        self.rect(x + third, by, third, band_h, style='F')
        # jaune
        self.set_fill_color(*YELLOW)
        self.rect(x + 2 * third, by, third, band_h, style='F')

        # textes année
        self.set_font('inter', 'B', 5)
        self.set_text_color(255)
        self.set_xy(x, by + 1)
        self.cell(third, band_h - 2, f"Année {annee}", align='C')
        self.set_text_color(*INK)
        self.set_xy(x + 2 * third, by + 1)
        self.cell(third, band_h - 2, f"Year {annee}", align='C')

        # étoile jaune au centre du bloc rouge
        self.set_text_color(*YELLOW)
        self.set_font('ZapfDingbats', '', 8)
        self.set_xy(x + third, by + 0.6)
        self.cell(third, band_h - 1.2, chr(72), align='C')

    # ------------------------------------------------------------------
    # Ajuste une chaîne à une largeur : on RÉDUIT d'abord la police (auto-shrink)
    # jusqu'à un plancher, pour garder le texte COMPLET (essentiel pour un nom).
    # Ce n'est qu'au plancher, si ça déborde encore (très rare), qu'on tronque.
    # Renvoie (texte, taille_de_police_à_utiliser).
    # ------------------------------------------------------------------
    def _fit_shrink(self, text, max_w, size, bold=False, min_size=5.0):
        if not text:
            return "", size
        family_style = 'B' if bold else ''
        s = size
        while s > min_size:
            self.set_font('inter', family_style, s)
            if self.get_string_width(text) <= max_w:
                return text, s  # tient complet à cette taille
            s -= 0.1
        # Au plancher : si ça déborde toujours, troncature ultime de sécurité.
        self.set_font('inter', family_style, min_size)
        if self.get_string_width(text) <= max_w:
            return text, min_size
        ell = "…"
        while text and self.get_string_width(text + ell) > max_w:
            text = text[:-1]
        return text + ell, min_size

    # ------------------------------------------------------------------
    # Réduit une chaîne pour qu'elle tienne dans une largeur donnée
    # (sinon fpdf déborde). Tronque avec "…" si nécessaire.
    # ------------------------------------------------------------------
    def _fit(self, text, max_w, size, bold=False):
        if not text:
            return ""
        self.set_font('inter', 'B' if bold else '', size)
        if self.get_string_width(text) <= max_w:
            return text
        ell = "…"
        while text and self.get_string_width(text + ell) > max_w:
            text = text[:-1]
        return text + ell


class StudentsIdentityCards(FPDF):

    def __init__(self, *args, **kwargs):
        super().__init__(orientation='P')
        self.add_font('inter', '', settings.INTER_REGULAR)
        self.add_font('inter', 'I', settings.INTER_ITALIC)
        self.add_font('inter', 'B', settings.INTER_BOLD)
        self.add_font('inter', 'BI', settings.INTER_BOLDITALIC)
        self.alias_nb_pages()
        self.set_margins(5, 5, 5)
        self.set_auto_page_break(auto=True, margin=12)
        self.set_font('inter', '', 8)
        self.data = kwargs.pop('data')
        logo, height = resize_image(self.data['school_data']['logo'], new_width=475, return_height=True) \
            if self.data['school_data']['logo'] != "static/image/no_image.jpg" else None
        height = int((height * 25.4) / 300) + 20 if height else 65
        height = height if self.data['school_data']['motto'] else None
        if 'students' in self.data.keys():
            # Cachet (rond établissement) ET visa (nominatif + signature) préchargés
            # EN BYTES une fois. On recrée un BytesIO neuf à chaque pose (jamais
            # épuisé, même en mode ZIP toutes les classes). Voir utils.stamp_bytes.
            self._cachet_bytes = stamp_bytes(self.data['school_data'].get('cachet'))  # noqa: F821
            self._visa_bytes = stamp_bytes(self.data['school_data'].get('visa'))  # noqa: F821
            self.id_cards(logo, height)
        else:
            self._cachet_bytes = None
            self._visa_bytes = None
            self.second_page()
            self.third_page((1, 2, 3))
            self.first_and_last_pages((1, 2, 3), logo, height)

    def id_cards(self, logo, height):
        students = self.data['students']
        is_not_modulo3 = len(students) % 3 != 0
        j = int(len(students) / 3)
        if is_not_modulo3:
            j += 1
        for i in range(j):
            if i == j -1 and is_not_modulo3:
                end = len(students)
            else:
                end = (i *3) + 3
            students_set = students[i * 3:end]
            self.add_page()
            x = 5
            y = 2.5
            for student in students_set:
                self.set_xy(x, y)
                self.line(x1=105, y1=y, x2=105, y2=y+90)
                table = Table(self, line_height=4, col_widths=(35, 60), text_align="L", first_row_as_headings=False,
                              borders_layout="NONE", markdown=True, align='L', width=96)
                row = table.row()
                row.cell("**Nom(s) :**")
                self.set_font_size(9)
                row.cell(f"**{student.nom}**", rowspan=2)
                row = table.row()
                self.set_font_size(7)
                row.cell("__Name(s)__")
                self.set_font_size(8)

                row = table.row()
                row.cell("**Préom(s) :**")
                self.set_font_size(9)
                row.cell(f"**{student.prenom if student.prenom else '/'}**", rowspan=2)
                row = table.row()
                self.set_font_size(7)
                row.cell("__Surname(s)__")
                self.set_font_size(8)

                row = table.row()
                row.cell("**Né(e) le :**")
                self.set_font_size(9)
                row.cell(f"**{format_date(student.date_naissance)}**", rowspan=2)
                row = table.row()
                self.set_font_size(7)
                row.cell("__Born on__")
                self.set_font_size(8)

                row = table.row()
                row.cell("**À :**")
                self.set_font_size(9)
                row.cell(f"**{student.lieu_naissance}**", rowspan=2)
                row = table.row()
                self.set_font_size(7)
                row.cell("__At__")
                self.set_font_size(8)

                row = table.row()
                row.cell("**Nationalité :**")
                self.set_font_size(9)
                row.cell("**Camerounaise**", rowspan=2)
                row = table.row()
                self.set_font_size(7)
                row.cell("__Nationality__")
                self.set_font_size(8)

                row = table.row()
                row.cell("**Fils ou Fille de :**")
                self.set_font_size(9)
                pere = f"**{student.pere.name}**" if student.pere else f"--{' ' * 65}--"
                row.cell(pere, rowspan=2)
                row = table.row()
                self.set_font_size(7)
                row.cell("__Son or Daughter of__")
                self.set_font_size(8)

                row = table.row()
                row.cell("**Et de :**")
                self.set_font_size(9)
                mere = f"**{student.mere.name}**" if student.mere else f"--{' ' * 65}--"
                row.cell(mere, rowspan=2)
                row = table.row()
                self.set_font_size(7)
                row.cell("__And of__")
                self.set_font_size(8)

                row = table.row()
                row.cell("**Adresse des Parents :**")
                self.set_font_size(9)
                contact = student.pere.contact if student.pere else ""
                if student.mere:
                    contact += f" / {student.mere.contact}" if contact else student.mere.contact
                contact = f"**{contact}**" if contact else f"--{' ' * 65}--"
                row.cell(contact, rowspan=2)
                row = table.row()
                self.set_font_size(7)
                row.cell("__Parent's adress__")
                self.set_font_size(8)

                row = table.row()
                row.cell("**Classe :**")
                self.set_font_size(9)
                row.cell(f"**{student.classe.code if student.classe else ' '}**", rowspan=2)
                row = table.row()
                self.set_font_size(7)
                row.cell("__Classroom__")
                self.set_font_size(8)

                row = table.row()
                row.cell("**N° Matricule :**")
                self.set_text_color(*RED)
                self.set_font_size(9)
                row.cell(f"**{student.unique_id if student.unique_id else ' '}**", rowspan=2)
                self.set_text_color(0)
                row = table.row()
                self.set_font_size(7)
                row.cell("__Register number__")
                self.set_font_size(8)

                row = table.row()
                row.cell("**Signature ou empreinte de l'Élève**\n", colspan=2, align='C')
                row = table.row()
                self.set_font_size(7)
                row.cell("__Student's signature__\n", colspan=2, align='C')
                self.set_font_size(8)

                table.render()
                self.dashed_line(x1=0, y1=y+92.5, x2=210, y2=y+92.5)
                y += 95
            self.third_page(students_set)
            self.first_and_last_pages(students_set, logo, height)

    def second_page(self):
        self.add_page()
        x = 5
        y = 2.5
        for _ in range(3):
            self.set_xy(x, y)
            self.line(x1=105, y1=y, x2=105, y2=y + 90)
            table = Table(self, line_height=4, col_widths=(35, 60), text_align="L", first_row_as_headings=False,
                          borders_layout="NONE", markdown=True, align='L', width=96)
            row = table.row()
            row.cell("**Nom(s) :**")
            self.set_font_size(9)
            row.cell(f"--{' ' * 65}--", rowspan=2)
            row = table.row()
            self.set_font_size(7)
            row.cell("__Name(s)__")
            self.set_font_size(8)

            row = table.row()
            row.cell("**Préom(s) :**")
            self.set_font_size(9)
            row.cell(f"--{' ' * 65}--", rowspan=2)
            row = table.row()
            self.set_font_size(7)
            row.cell("__Surname(s)__")
            self.set_font_size(8)

            row = table.row()
            row.cell("**Né(e) le :**")
            self.set_font_size(9)
            row.cell(f"--{' ' * 65}--", rowspan=2)
            row = table.row()
            self.set_font_size(7)
            row.cell("__Born on__")
            self.set_font_size(8)

            row = table.row()
            row.cell("**À :**")
            self.set_font_size(9)
            row.cell(f"--{' ' * 65}--", rowspan=2)
            row = table.row()
            self.set_font_size(7)
            row.cell("__At__")
            self.set_font_size(8)

            row = table.row()
            row.cell("**Nationalité :**")
            self.set_font_size(9)
            row.cell(f"--{' ' * 65}--", rowspan=2)
            row = table.row()
            self.set_font_size(7)
            row.cell("__Nationality__")
            self.set_font_size(8)

            row = table.row()
            row.cell("**Fils ou Fille de :**")
            self.set_font_size(9)
            row.cell(f"--{' ' * 65}--", rowspan=2)
            row = table.row()
            self.set_font_size(7)
            row.cell("__Son or Daughter of__")
            self.set_font_size(8)

            row = table.row()
            row.cell("**Et de :**")
            self.set_font_size(9)
            row.cell(f"--{' ' * 65}--", rowspan=2)
            row = table.row()
            self.set_font_size(7)
            row.cell("__And of__")
            self.set_font_size(8)

            row = table.row()
            row.cell("**Adresse des Parents :**")
            self.set_font_size(9)
            row.cell(f"--{' ' * 65}--", rowspan=2)
            row = table.row()
            self.set_font_size(7)
            row.cell("__Parent's adress__")
            self.set_font_size(8)

            row = table.row()
            row.cell("**Classe :**")
            self.set_font_size(9)
            row.cell(f"--{' ' * 65}--", rowspan=2)
            row = table.row()
            self.set_font_size(7)
            row.cell("__Classroom__")
            self.set_font_size(8)

            row = table.row()
            row.cell("**N° Matricule :**")
            self.set_font_size(9)
            row.cell(f"--{' ' * 65}--", rowspan=2)
            row = table.row()
            self.set_font_size(7)
            row.cell("__Register number__")
            self.set_font_size(8)

            row = table.row()
            row.cell("**Signature ou empreinte de l'Élève**\n", colspan=2, align='C')
            row = table.row()
            self.set_font_size(7)
            row.cell("__Student's signature__\n", colspan=2, align='C')
            self.set_font_size(8)

            table.render()
            self.dashed_line(x1=0, y1=y + 92.5, x2=210, y2=y + 92.5)
            y += 95

    def third_page(self, elts):
        x, y = 110, 2.5
        for elt in elts:
            photo = "" if isinstance(elt, int) else elt.photo
            if photo:
                self.image(resize_image(photo, id_card=True), x=140, y=y, w=35, h=40)
            else:
                self.set_xy(140, y)
                self.cell(w=35, h=40, text="__Photo__", border=True, align='C', markdown=True)
            self.set_xy(x, y+44)
            self.set_font_size(9)
            self.cell(w=95, h=3, text=f"**Fait à {self.data['school_data']['localite']}, le **--{' ' * 20}--", align='C', markdown=True)
            self.set_xy(x, y + 49)
            self.cell(w=95, h=4, text=f"**LE {self.data['school_data']['chef'].upper()}**", align='C', markdown=True)
            self.set_font_size(8)
            self.set_xy(x, y + 53)
            self.cell(w=95, h=3, text=f"__The Principal", align='C', markdown=True)
            # --- Cachet (rond établissement) et Visa (nominatif + signature) ---
            paste_stamp(self, self._cachet_bytes,  # noqa: F821
                        x=x + 3, y=y + 32, w=40)
            paste_stamp(self, self._visa_bytes,  # noqa: F821
                        x=x + 45, y=y + 56, w=45)
            if self._visa_bytes and self._cachet_bytes:
                l = self.get_string_width(f"**Fait à {self.data['school_data']['localite']}, le ",
                                      markdown=True)
                self.set_xy(x + 39 + (l / 2), y + 44)
                self.set_font_size(9)
                self.set_text_color(*RED)
                from babel.dates import format_date as babel_format_date
                self.cell(w=18, h=3, text=f"**{babel_format_date(datetime.now(), format="short", locale="fr_FR")}**",
                          align='L', markdown=True)
                self.set_text_color(0)
            self.set_font_size(7)
            self.line(x1=110, y1=y+77.5, x2=205, y2=y+77.5)
            self.set_xy(x, y + 78)
            self.cell(w=95, h=3, text="CETTE CARTE EST STRICTEMENT INDIVIDUELLE ET PERSONNELLE", align='C')
            self.set_xy(x, y + 81)
            self.cell(w=95, h=3, text="ELLE NE PEUT EN AUCUN CAS ËTRE PRËTÉE OU CÉDÉE", align='C')
            self.set_xy(x, y + 86)
            self.cell(w=95, h=3, text="__THIS CARD IS STRICTLY INDIVIDUAL AND PERSONNAL__", align='C', markdown=True)
            self.set_xy(x, y + 89)
            self.cell(w=95, h=3, text="__IT CAN NOT BE RENT OR GIVEN OUT__", align='C', markdown=True)
            self.set_font_size(8)

            y += 95

    def drapeu(self, x, y, w=20):
        with self.local_context(fill_opacity=0.8):
            self.image("static/image/drapeau.jpg", x=x, y=y, w=w, keep_aspect_ratio=True)

    def first_and_last_pages(self, elts, logo, height):
        self.add_page()
        x, y, yi, yd, ym = 110, 2.5, 15, 2.5, height
        for elt in elts:
            if logo:
                self.image(logo, x=32.5, y=yi, w=40, keep_aspect_ratio=True)
                yi += 95
            if height:
                self.set_xy(5, ym)
                self.set_font_size(10)
                self.cell(w=95, h=10, text=f"**__{self.data['school_data']['motto']}__**", markdown=True, align='C')
                self.set_font_size(8)
                ym += 95

            self.set_xy(x, y)
            self.drapeu(x=108, y=yd)
            self.line(x1=105, y1=y, x2=105, y2=y + 90)

            self.cell(w=47.5, h=4, text="**RÉPUBLIQUE DU CAMEROUN**", align='C', markdown=True)
            self.cell(w=47.5, h=4, text="**REPUBLIC OF CAMEROON**", align='C', markdown=True)
            self.set_font_size(7)
            self.set_xy(110, y+4)
            self.cell(w=47.5, h=4, text="__Paix - Travail - Patrie__", align='C', markdown=True)
            self.cell(w=47.5, h=4, text="__Peace - Work - Fatherland__", align='C', markdown=True)
            self.set_font_size(6)
            self.set_xy(110, y+8)
            self.cell(w=47.5, h=4, text="*********", align='C', markdown=True)
            self.cell(w=47.5, h=4, text="*********", align='C', markdown=True)
            self.set_font_size(8)
            self.set_xy(110, y+12)
            self.cell(w=95, h=4, text="**MINISTÈRE DES ENSEIGNEMENTS SECONDAIRES**", align='C', markdown=True)
            self.set_font_size(7)
            self.set_xy(110, y+16)
            self.cell(w=95, h=4, text="__MINISTRY OF SECONDARY EDUCATION__", align='C', markdown=True)
            self.set_font_size(6)
            self.set_xy(110, y+20)
            self.cell(w=95, h=4, text="************", align='C', markdown=True)
            self.set_font_size(8)
            self.set_xy(110, y+24)
            table = Table(self, line_height=4, col_widths=(1, 1), text_align="R", first_row_as_headings=False, cell_fill_color=None,
                          borders_layout="NONE", markdown=True, align='L', width=95, cell_fill_mode=TableCellFillMode.NONE)
            row = table.row()
            row.cell(f"**{self.data['school_data']['region']}**", align='C', colspan=2)
            self.set_font_size(7)
            row = table.row()
            row.cell(f"__{self.data['school_data']['rgn']}__", align='C', colspan=2)
            self.set_font_size(6)
            row = table.row()
            row.cell("************", align='C', v_align=VAlign.T, colspan=2)
            self.set_font_size(8)

            row = table.row()
            row.cell(f"**{self.data['school_data']['departement']}**", align='C', colspan=2)
            self.set_font_size(7)
            row = table.row()
            row.cell(f"__{self.data['school_data']['dptm']}__", align='C', colspan=2)
            self.set_font_size(6)
            row = table.row()
            row.cell("************", align='C', v_align=VAlign.T, colspan=2)
            row = table.row()
            self.set_fill_color(75)
            self.set_text_color(255)
            row.cell(" ", colspan=2)
            self.set_font_size(12)

            row = table.row()
            nom = f"**{self.data['school_data']['nom']}**"
            row.cell(nom, align='C', colspan=2)
            if self.get_string_width(nom, markdown=True) < 96:
                self.set_font_size(10)
                row = table.row()
                row.cell(f"__{(self.data['school_data']['name'])}__", align='C', colspan=2)
            self.set_font_size(8)
            row = table.row()
            row.cell(f"B.P. {self.data['school_data']['po_box']}", align='C', v_align=VAlign.T, colspan=2)
            self.set_text_color(0)
            self.set_fill_color(255)
            row = table.row()
            row.cell(" ", colspan=2)

            self.set_font_size(9)
            row = table.row()
            row.cell("**CARTE D'IDENTITÉ SCOLAIRE**", align='C', colspan=2)
            self.set_font_size(8)
            row = table.row()
            row.cell("__STUDENT IDENTITY CARD__", align='C', v_align=VAlign.T, colspan=2)
            row = table.row()
            number = f"--{' ' * 20}--" if isinstance(elt, int) else f"**{self.data['annee'][-2:]}{elt.csi_number}**"
            row.cell("**N°**", align='R')
            if not isinstance(elt, int):
                self.set_text_color(*RED)
            row.cell(number, align='L')
            self.set_text_color(0)

            row = table.row()
            row.cell(" ", colspan=2)
            row = table.row()
            row.cell("**ANNÉE SCOLAIRE**", align='R')
            row.cell(f"**{self.data['annee']}**", align='L')
            row = table.row()
            self.set_font_size(7)
            row.cell("__SCHOOL YEAR__", align='R')
            self.set_font_size(8)

            table.render()
            self.dashed_line(x1=0, y1=y + 92.5, x2=210, y2=y + 92.5)
            y += 95
            yd += 95
